from typing import Optional, Sequence
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

DASH = "-   "

# Sheet 6 row order: identical to Sheet 5 but WITHOUT the per-year rows
ROW_ORDER_S6: Sequence[str] = (
    "Loan amounts (USD)",
    "Sum insured",
    "Area",
    "Region",
    # (no years)
    "Average Payout",
    "SD",
    "CoV",
    "Min",
    "Max",
    "90th percentile",
    "95th percentile",
    "Number of Pixels",
    "Number of Zero and Blank Pixels",
    "Number of Blank Pixels",
    "Number of Zero Pixel",
    "Average non-zero/blank pixel CoV",
)

AREA_COLORS_HEX = {
    "Northern Zone": "1F77B4",
    "Central Zone": "2CA02C",
    "Lake Zone": "FF7F0E",
    "Western Zone": "9467BD",
    "Southern Highlands Zone": "8C564B",
    "Coastal Zone": "17BECF",
    "Zanzibar (Islands)": "7F7F7F",
}

def _area_fill(area_name: str) -> Optional[PatternFill]:
    if not area_name:
        return None
    for k, hexv in AREA_COLORS_HEX.items():
        if str(area_name).strip().lower() == k.lower():
            return PatternFill(fill_type="solid", start_color=f"FF{hexv}", end_color=f"FF{hexv}")
    return None

def _auto_size(ws: Worksheet, min_width=6, max_width=22):
    for col in range(1, ws.max_column + 1):
        mx = 0
        for vals in ws.iter_cols(min_col=col, max_col=col, min_row=1, max_row=ws.max_row, values_only=True):
            for v in vals:
                if v is None:
                    continue
                mx = max(mx, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, min(max_width, mx + 2))

def _detect_year_rows(df_numeric: pd.DataFrame):
    years = []
    for idx in df_numeric.index:
        s = str(idx)
        if s.isdigit():
            y = int(s)
            if 1900 <= y <= 2100:
                years.append(s)
    return sorted(years, key=int)

def _xq(text: str) -> str:
    """Excel-safe double-quoted literal."""
    if text is None:
        text = ""
    return '"' + str(text).replace('"', '""') + '"'

def build_excel_sheet6(
    df_wide_numeric: pd.DataFrame,
    df_wide_formatted: pd.DataFrame,
    wb: Optional[Workbook] = None,
    sheet_name: str = "6. Regional Totals (Stats Only)",
    sheet5_name: str = "5. Regional Totals",   # where the year rows live
) -> Workbook:
    # sanity
    if not df_wide_numeric.index.equals(df_wide_formatted.index):
        raise ValueError("Index mismatch between numeric and formatted wide dataframes.")
    if list(df_wide_numeric.columns) != list(df_wide_formatted.columns):
        raise ValueError("Column mismatch between numeric and formatted wide dataframes.")

    # detect year rows to know which rows to reference on Sheet 5
    year_rows = _detect_year_rows(df_wide_numeric)
    if not year_rows:
        # It's OK if there are no year rows; stats will gracefully blank via COUNT guards.
        pass

    # Compute the row numbers where those year totals appear on Sheet 5
    # Sheet 5 inserts years BEFORE "Average Payout"
    ROW_ORDER_BASE_S5 = (
        "Loan amounts (USD)","Sum insured","Area","Region",
        # years are inserted here on sheet 5
        "Average Payout","SD","CoV","Min","Max","90th percentile","95th percentile",
        "Number of Pixels","Number of Zero and Blank Pixels","Number of Blank Pixels",
        "Number of Zero Pixel","Average non-zero/blank pixel CoV",
    )
    ordered_rows_s5 = []
    for name in ROW_ORDER_BASE_S5:
        if name == "Average Payout":
            ordered_rows_s5.extend(year_rows)
        ordered_rows_s5.append(name)

    first_year_row_s5 = (ordered_rows_s5.index(year_rows[0]) + 1) if year_rows else None
    last_year_row_s5  = (ordered_rows_s5.index(year_rows[-1]) + 1) if year_rows else None

    wb = wb or Workbook()
    if (wb.active and wb.active.max_row == 1 and wb.active.max_column == 1
            and wb.active.title.lower().startswith("sheet")):
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)

    bold  = Font(bold=True)
    left  = Alignment(horizontal="left",  vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    center= Alignment(horizontal="center",vertical="center")

    # Labels in column A
    NOT_BOLD = {"Number of Zero and Blank Pixels", "Number of Blank Pixels",
                "Number of Zero Pixel", "Average non-zero/blank pixel CoV"}
    for r_idx, label in enumerate(ROW_ORDER_S6, start=1):
        c = ws.cell(row=r_idx, column=1, value=label)
        if label not in NOT_BOLD:
            c.font = bold
        c.alignment = left

    # Pointers to Sheets 3 & 4 (dynamic horizontal width)
    HCOUNT = "COUNTA('3. Payout Amounts'!F9:XFD9)"

    def row_range_on_sheet3(row_num: int) -> str:
        # 1×N horizontal range starting at F<row>, width = HCOUNT
        return f"OFFSET('3. Payout Amounts'!F{row_num},0,0,1,{HCOUNT})"

    REGION_ROW    = row_range_on_sheet3(6)   # '3. Payout Amounts' row 6
    AREA_ROW      = row_range_on_sheet3(5)   # row 5
    LOAN_ROW      = row_range_on_sheet3(3)   # row 3
    SUMINS_ROW    = row_range_on_sheet3(4)   # row 4
    AVGPIX_ROW4   = f"OFFSET('4. Pixel Stats'!F10,0,0,1,{HCOUNT})"  # per-pixel avg payouts (USD)
    COVPIX_ROW4   = f"OFFSET('4. Pixel Stats'!F12,0,0,1,{HCOUNT})"  # per-pixel CoV

    # Column loop (B → …)
    for col_idx, col_name in enumerate(df_wide_numeric.columns, start=2):
        colL = get_column_letter(col_idx)

        area_val   = str(df_wide_formatted.at["Area",   col_name]) if "Area"   in df_wide_formatted.index else ""
        region_val = str(df_wide_formatted.at["Region", col_name]) if "Region" in df_wide_formatted.index else ""

        is_overall       = (region_val.strip().lower() == "overall total")
        is_total_of_area = (region_val.strip().lower() == "total" and area_val.strip() and not is_overall)

        # Row indices (constant for this sheet)
        r_loan = ROW_ORDER_S6.index("Loan amounts (USD)") + 1
        r_si   = ROW_ORDER_S6.index("Sum insured") + 1
        r_area = ROW_ORDER_S6.index("Area") + 1
        r_reg  = ROW_ORDER_S6.index("Region") + 1

        # ---- Area/Region with special handling for Overall Total (single bold merged cell) ----
        if is_overall:
            merged_cell = ws.cell(row=r_area, column=col_idx)
            merged_cell.value = "Overall Total"
            merged_cell.font = bold
            merged_cell.alignment = center
            ws.cell(row=r_reg, column=col_idx).value = None
            ws.merge_cells(start_row=r_area, start_column=col_idx, end_row=r_reg, end_column=col_idx)
        else:
            ca = ws.cell(row=r_area, column=col_idx)
            ca.value = area_val if area_val and area_val != "-" else DASH
            fill = _area_fill(area_val)
            if fill:
                ca.fill = fill
            ca.alignment = center

            cr = ws.cell(row=r_reg, column=col_idx)
            cr.value = region_val if region_val and region_val != "-" else DASH
            cr.alignment = center

        # ---- Loan amounts (USD) ----
        cl = ws.cell(row=r_loan, column=col_idx)
        if is_overall:
            cl.value = f"=SUM({LOAN_ROW})"
        elif is_total_of_area:
            cl.value = f"=SUMPRODUCT(--({AREA_ROW}={_xq(area_val)}),{LOAN_ROW})"
        else:
            cl.value = f"=SUMPRODUCT(--({REGION_ROW}={_xq(region_val)}),{LOAN_ROW})"
        cl.number_format = "# ##0"
        cl.alignment = right

        # ---- Sum insured ----
        cs = ws.cell(row=r_si, column=col_idx)
        if is_overall:
            cs.value = f"=SUM({SUMINS_ROW})"
        elif is_total_of_area:
            cs.value = f"=SUMPRODUCT(--({AREA_ROW}={_xq(area_val)}),{SUMINS_ROW})"
        else:
            cs.value = f"=SUMPRODUCT(--({REGION_ROW}={_xq(region_val)}),{SUMINS_ROW})"
        cs.number_format = "# ##0"
        cs.alignment = right

        # ---- Statistics that reference the year totals on Sheet 5 ----
        # (We compute stats over '5. Regional Totals' year rows for the same column)
        def years_col_range_on_sheet5() -> str:
            if not year_rows:
                # harmless single-cell range ensures COUNT()=0 and stats blank out
                return f"'{sheet5_name}'!{colL}1:{colL}1"
            return f"'{sheet5_name}'!{colL}{first_year_row_s5}:{colL}{last_year_row_s5}"

        # Average
        r_avg = ROW_ORDER_S6.index("Average Payout") + 1
        ws.cell(row=r_avg, column=col_idx).value = (
            f"=IF(COUNT({years_col_range_on_sheet5()})=0,\"\",AVERAGE({years_col_range_on_sheet5()}))"
        )
        ws.cell(row=r_avg, column=col_idx).number_format = "# ##0"
        ws.cell(row=r_avg, column=col_idx).alignment = right

        # SD
        r_sd = ROW_ORDER_S6.index("SD") + 1
        ws.cell(row=r_sd, column=col_idx).value = (
            f"=IF(COUNT({years_col_range_on_sheet5()})<=1,\"\",STDEV({years_col_range_on_sheet5()}))"
        )
        ws.cell(row=r_sd, column=col_idx).number_format = "# ##0"
        ws.cell(row=r_sd, column=col_idx).alignment = right

        # CoV = SD / Average (blank-safe)
        r_cov = ROW_ORDER_S6.index("CoV") + 1
        avg_ref = f"{colL}{r_avg}"; sd_ref = f"{colL}{r_sd}"
        ws.cell(row=r_cov, column=col_idx).value = f"=IF(OR(ISBLANK({avg_ref}),{avg_ref}=0,ISBLANK({sd_ref})),\"\",{sd_ref}/{avg_ref})"
        ws.cell(row=r_cov, column=col_idx).number_format = "0.00"
        ws.cell(row=r_cov, column=col_idx).alignment = right

        # Min / Max / P90 / P95 over the year block on Sheet 5
        r_min = ROW_ORDER_S6.index("Min") + 1
        ws.cell(row=r_min, column=col_idx).value = (
            f"=IF(COUNT({years_col_range_on_sheet5()})=0,\"\",MIN({years_col_range_on_sheet5()}))"
        )
        ws.cell(row=r_min, column=col_idx).number_format = "# ##0"
        ws.cell(row=r_min, column=col_idx).alignment = right

        r_max = ROW_ORDER_S6.index("Max") + 1
        ws.cell(row=r_max, column=col_idx).value = (
            f"=IF(COUNT({years_col_range_on_sheet5()})=0,\"\",MAX({years_col_range_on_sheet5()}))"
        )
        ws.cell(row=r_max, column=col_idx).number_format = "# ##0"
        ws.cell(row=r_max, column=col_idx).alignment = right

        r_p90 = ROW_ORDER_S6.index("90th percentile") + 1
        ws.cell(row=r_p90, column=col_idx).value = (
            f"=IF(COUNT({years_col_range_on_sheet5()})=0,\"\",PERCENTILE({years_col_range_on_sheet5()},0.9))"
        )
        ws.cell(row=r_p90, column=col_idx).number_format = "# ##0"
        ws.cell(row=r_p90, column=col_idx).alignment = right

        r_p95 = ROW_ORDER_S6.index("95th percentile") + 1
        ws.cell(row=r_p95, column=col_idx).value = (
            f"=IF(COUNT({years_col_range_on_sheet5()})=0,\"\",PERCENTILE({years_col_range_on_sheet5()},0.95))"
        )
        ws.cell(row=r_p95, column=col_idx).number_format = "# ##0"
        ws.cell(row=r_p95, column=col_idx).alignment = right

        # ---- Counts (same definitions as Sheet 5) ----
        r_np = ROW_ORDER_S6.index("Number of Pixels") + 1
        if is_overall:
            ws.cell(row=r_np, column=col_idx).value = f"={HCOUNT}"
        elif is_total_of_area:
            ws.cell(row=r_np, column=col_idx).value = f"=SUMPRODUCT(--({AREA_ROW}={_xq(area_val)}))"
        else:
            ws.cell(row=r_np, column=col_idx).value = f"=SUMPRODUCT(--({REGION_ROW}={_xq(region_val)}))"
        ws.cell(row=r_np, column=col_idx).number_format = "# ##0"
        ws.cell(row=r_np, column=col_idx).alignment = right

        r_nb = ROW_ORDER_S6.index("Number of Blank Pixels") + 1
        if is_overall:
            blank_formula = f"=SUMPRODUCT(--(LEN({AVGPIX_ROW4})=0))"
        elif is_total_of_area:
            blank_formula = f"=SUMPRODUCT(--({AREA_ROW}={_xq(area_val)}),--(LEN({AVGPIX_ROW4})=0))"
        else:
            blank_formula = f"=SUMPRODUCT(--({REGION_ROW}={_xq(region_val)}),--(LEN({AVGPIX_ROW4})=0))"
        ws.cell(row=r_nb, column=col_idx).value = blank_formula
        ws.cell(row=r_nb, column=col_idx).number_format = "# ##0"
        ws.cell(row=r_nb, column=col_idx).alignment = right

        r_nz = ROW_ORDER_S6.index("Number of Zero Pixel") + 1
        if is_overall:
            zero_formula = f"=SUMPRODUCT(--(LEN({AVGPIX_ROW4})<>0),--({AVGPIX_ROW4}=0))"
        elif is_total_of_area:
            zero_formula = f"=SUMPRODUCT(--({AREA_ROW}={_xq(area_val)}),--(LEN({AVGPIX_ROW4})<>0),--({AVGPIX_ROW4}=0))"
        else:
            zero_formula = f"=SUMPRODUCT(--({REGION_ROW}={_xq(region_val)}),--(LEN({AVGPIX_ROW4})<>0),--({AVGPIX_ROW4}=0))"
        ws.cell(row=r_nz, column=col_idx).value = zero_formula
        ws.cell(row=r_nz, column=col_idx).number_format = "# ##0"
        ws.cell(row=r_nz, column=col_idx).alignment = right

        r_nzb = ROW_ORDER_S6.index("Number of Zero and Blank Pixels") + 1
        ws.cell(row=r_nzb, column=col_idx).value = f"={get_column_letter(col_idx)}{r_nb}+{get_column_letter(col_idx)}{r_nz}"
        ws.cell(row=r_nzb, column=col_idx).number_format = "# ##0"
        ws.cell(row=r_nzb, column=col_idx).alignment = right

        r_cov2 = ROW_ORDER_S6.index("Average non-zero/blank pixel CoV") + 1
        if is_overall:
            num = f"SUMPRODUCT(--(LEN({AVGPIX_ROW4})<>0),--({AVGPIX_ROW4}<>0),{COVPIX_ROW4})"
            den = f"SUMPRODUCT(--(LEN({AVGPIX_ROW4})<>0),--({AVGPIX_ROW4}<>0))"
        elif is_total_of_area:
            num = f"SUMPRODUCT(--({AREA_ROW}={_xq(area_val)}),--(LEN({AVGPIX_ROW4})<>0),--({AVGPIX_ROW4}<>0),{COVPIX_ROW4})"
            den = f"SUMPRODUCT(--({AREA_ROW}={_xq(area_val)}),--(LEN({AVGPIX_ROW4})<>0),--({AVGPIX_ROW4}<>0))"
        else:
            num = f"SUMPRODUCT(--({REGION_ROW}={_xq(region_val)}),--(LEN({AVGPIX_ROW4})<>0),--({AVGPIX_ROW4}<>0),{COVPIX_ROW4})"
            den = f"SUMPRODUCT(--({REGION_ROW}={_xq(region_val)}),--(LEN({AVGPIX_ROW4})<>0),--({AVGPIX_ROW4}<>0))"
        ws.cell(row=r_cov2, column=col_idx).value = f"=IF({den}=0,\"\",{num}/{den})"
        ws.cell(row=r_cov2, column=col_idx).number_format = "0.00"
        ws.cell(row=r_cov2, column=col_idx).alignment = right

    # Freeze panes (below Region, after first column)
    ws.freeze_panes = ws.cell(row=ROW_ORDER_S6.index("Region") + 2, column=2)

    # Column A wider; auto-size others
    ws.column_dimensions['A'].width = 36
    _auto_size(ws, max_width=22)



    # === BEGIN: Formatting tweaks per request (v2) ===

    from openpyxl.utils import get_column_letter as _gcl6

    # Task 1: Set data columns (B and onward) width -> 21.4

    for _c in range(2, ws.max_column + 1):

        ws.column_dimensions[_gcl6(_c)].width = 21.4

    # === END: Formatting tweaks per request (v2) ===


    return wb
