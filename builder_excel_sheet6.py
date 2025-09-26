from typing import Optional, Sequence, Union, Dict
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

DASH = "-   "   # placeholder (dash + spaces like your sample)

# Display order for Sheet 6
ROW_DISPLAY_ORDER: Sequence[str] = (
    "Loan amounts (USD)",
    "Sum insured",
    "Area",
    "Region",
    "Average Payout",
    "SD",
    "CoV",  # corresponds to 'Area CoV' in numeric source
    "Min",
    "Max",
    "90th percentile",
    "95th percentile",
    "Average non-zero/blank pixel CoV",
)

# Map display labels to the source row keys in df_wide_numeric/df_wide_formatted
# 'CoV' comes from 'Area CoV'
ROW_SOURCE_BY_DISPLAY: Dict[str, str] = {
    "Loan amounts (USD)": "Loan amounts (USD)",
    "Sum insured": "Sum insured",
    "Area": "Area",
    "Region": "Region",
    "Average Payout": "Average Payout",
    "SD": "SD",
    "CoV": "Area CoV",  # important mapping
    "Min": "Min",
    "Max": "Max",
    "90th percentile": "90th percentile",
    "95th percentile": "95th percentile",
    "Average non-zero/blank pixel CoV": "Average non-zero/blank pixel CoV",
}

# Rows to render as integers (with thousands separator and dash for 0/NaN)
INT_ROWS = {
    "Loan amounts (USD)",
    "Sum insured",
    "Average Payout",
    "SD",
    "Min",
    "Max",
    "90th percentile",
    "95th percentile",
}

# Rows to render as CoV (two decimals, comma decimal separator)
COV_ROWS = {
    "CoV",
    "Average non-zero/blank pixel CoV",
}

AREA_COLORS_HEX = {
    "Northern Zone": "1F77B4",
    "Central Zone": "2CA02C",
    "Lake Zone": "FF7F0E",
    "Western Zone": "9467BD",
    "Southern Highlands Zone": "8C564B",
    "Coastal Zone": "17BECF",
    "Zanzibar (Islands)": "7F7F7F",
}

def _format_cov(x: Optional[float]) -> str:
    if x is None or pd.isna(x):
        return DASH
    return f"{float(x):.2f}".replace(".", ",")

def _auto_size(ws: Worksheet, min_width=6, max_width=32):
    for col in range(1, ws.max_column + 1):
        mx = 0
        for cell in ws.iter_cols(min_col=col, max_col=col,
                                 min_row=1, max_row=ws.max_row, values_only=True):
            for v in cell:
                if v is None:
                    continue
                ln = len(str(v))
                if ln > mx:
                    mx = ln
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, min(max_width, mx + 2))

def _area_fill(area_name: str) -> Optional[PatternFill]:
    if not area_name:
        return None
    for k, hexv in AREA_COLORS_HEX.items():
        if str(area_name).strip().lower() == k.lower():
            return PatternFill(fill_type="solid", start_color=f"FF{hexv}", end_color=f"FF{hexv}")
    return None

def build_excel_sheet6(
    df_wide_numeric: pd.DataFrame,
    df_wide_formatted: pd.DataFrame,
    wb: Optional[Workbook] = None,
    sheet_name: str = "6. Statistics",
    merge_overall_total: bool = True,
) -> Workbook:
    # Validate shape and indexes/columns
    if not df_wide_numeric.index.equals(df_wide_formatted.index):
        raise ValueError("Index mismatch between numeric and formatted wide dataframes.")
    if list(df_wide_numeric.columns) != list(df_wide_formatted.columns):
        raise ValueError("Column mismatch between numeric and formatted wide dataframes.")

    # Create or reuse sheet
    wb = wb or Workbook()
    if (wb.active and wb.active.max_row == 1 and wb.active.max_column == 1
            and wb.active.title.lower().startswith("sheet")):
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)

    bold = Font(bold=True)
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    # Rows whose label (first column) should NOT be bold (area/region only)
    NOT_BOLD_ROWS = {"Area", "Region"}

    # Write table row by row
    for row_idx_excel, display_label in enumerate(ROW_DISPLAY_ORDER, start=1):
        source_key = ROW_SOURCE_BY_DISPLAY.get(display_label, display_label)

        label_cell = ws.cell(row=row_idx_excel, column=1, value=display_label)
        if display_label not in NOT_BOLD_ROWS:
            label_cell.font = bold
        label_cell.alignment = left

        for col_offset, col_label in enumerate(df_wide_numeric.columns, start=2):
            # Use source key to fetch values; if missing, treat as None
            raw_val = (
                df_wide_numeric.at[source_key, col_label]
                if source_key in df_wide_numeric.index else None
            )
            excel_cell = ws.cell(row=row_idx_excel, column=col_offset)

            if display_label == "Area":
                # Prefer formatted for clean text
                val = (
                    df_wide_formatted.at[source_key, col_label]
                    if source_key in df_wide_formatted.index else None
                )
                if val is None or pd.isna(val) or str(val).strip() == "":
                    excel_cell.value = DASH
                else:
                    excel_cell.value = val
                    fill = _area_fill(str(val))
                    if fill:
                        excel_cell.fill = fill
                excel_cell.alignment = center
                continue

            if display_label == "Region":
                val = (
                    df_wide_formatted.at[source_key, col_label]
                    if source_key in df_wide_formatted.index else None
                )
                excel_cell.value = DASH if val is None or pd.isna(val) or str(val).strip() == "" else val
                excel_cell.alignment = center
                continue

            if display_label in COV_ROWS:
                excel_cell.value = _format_cov(raw_val)
                excel_cell.alignment = right
                continue

            if display_label in INT_ROWS:
                if raw_val is None or pd.isna(raw_val) or (
                    isinstance(raw_val, (int, float)) and float(raw_val) == 0
                ):
                    excel_cell.value = DASH
                else:
                    try:
                        ival = int(round(float(raw_val)))
                        excel_cell.value = ival
                        excel_cell.number_format = "# ##0"
                    except Exception:
                        excel_cell.value = DASH
                excel_cell.alignment = right
                continue

            # Fallback as text
            excel_cell.value = DASH if raw_val is None or pd.isna(raw_val) else str(raw_val)
            excel_cell.alignment = right

    # Merge "Overall Total" vertically across Area + Region columns if present
    if merge_overall_total and "Area" in ROW_DISPLAY_ORDER and "Region" in ROW_DISPLAY_ORDER:
        area_row_num = ROW_DISPLAY_ORDER.index("Area") + 1
        region_row_num = ROW_DISPLAY_ORDER.index("Region") + 1
        for col_offset, col_label in enumerate(df_wide_numeric.columns, start=2):
            area_cell_val = ws.cell(row=area_row_num, column=col_offset).value
            if area_cell_val and str(area_cell_val).strip().lower() == "overall total":
                ws.cell(row=region_row_num, column=col_offset).value = None
                ws.merge_cells(start_row=area_row_num, start_column=col_offset,
                               end_row=region_row_num, end_column=col_offset)
                merged = ws.cell(row=area_row_num, column=col_offset)
                merged.alignment = center
                merged.font = bold
                break

    # Freeze panes below Region row, after first column
    if "Region" in ROW_DISPLAY_ORDER:
        freeze_row = ROW_DISPLAY_ORDER.index("Region") + 2
        ws.freeze_panes = ws.cell(row=freeze_row, column=2)

    # Column sizing
    ws.column_dimensions[get_column_letter(1)].width = 36
    _auto_size(ws, max_width=22)

    return wb

def save_sheet6_from_statistics(
    df_wide_numeric: pd.DataFrame,
    df_wide_formatted: pd.DataFrame,
    file_path: Union[str, Path],
    sheet_name: str = "6. Statistics",
) -> str:
    """
    Build '6. Statistics' from provided wide dataframes and save to file_path.
    - df_wide_numeric: numeric wide dataframe containing rows keyed by ROW_SOURCE_BY_DISPLAY values
    - df_wide_formatted: formatted wide dataframe aligned with df_wide_numeric
    """
    p = Path(file_path)
    p.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb = build_excel_sheet6(df_wide_numeric, df_wide_formatted, wb=wb, sheet_name=sheet_name)
    wb.save(p)
    return str(p)

if __name__ == "__main__":
    print("Module ready: build_excel_sheet6 with 'CoV' mapped to 'Area CoV'.")