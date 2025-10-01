import pandas as pd
from typing import Optional, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

def _norm(s: str) -> str:
    return str(s).strip().lower().replace(" ", "").replace("_", "")

def _resolve_cols(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    norm_to_orig = {_norm(c): c for c in df.columns}
    def pick(*aliases: str) -> Optional[str]:
        for a in aliases:
            k = _norm(a)
            if k in norm_to_orig:
                return norm_to_orig[k]
        return None
    return {
        "pixel_key": pick("Pixel_ID", "pixelid", "pixel"),
        "pixel_id":  pick("Pixel_ID", "pixelid"),
        "year":      pick("Year", "year"),
        "area":      pick("Area", "area_ha", "hectares"),
    }

def _first_non_null(s: pd.Series):
    s2 = s.dropna()
    return s2.iloc[0] if not s2.empty else None

def _fill(hex6: str) -> PatternFill:
    return PatternFill(fill_type="solid", start_color=f"FF{hex6}", end_color=f"FF{hex6}")

def _autosize(ws, c1: int, c2: int, min_w: int = 8, max_w: int = 40):
    for col in range(c1, c2 + 1):
        m = 0
        for vals in ws.iter_cols(min_col=col, max_col=col, min_row=1, max_row=ws.max_row, values_only=True):
            for v in vals:
                if v is None:
                    continue
                m = max(m, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = max(min_w, min(max_w, m + 2))

def ws_title_is_default(title: str) -> bool:
    return str(title).lower().startswith("sheet")

def build_excel_sheet4(
    df: pd.DataFrame,
    wb: Optional[Workbook] = None,
    sheet_name: str = "4. Pixel Stats"
) -> Workbook:
    """
    Sheet 4: Pixel-by-pixel stats (16-row layout; headers not dynamic; Pixel count = values 1..N)

    Row labels in column E; pixel columns F→:
      1  Pixel count
      2  Loan amounts (USD)
      3  Sum insured
      4  Area   (color-filled like Sheet 3)
      5  Region
      6  Pixel Lon
      7  Pixel Lat
      8  Pixel ID
      9  Total   (INTENTIONALLY EMPTY)
     10  Average Payout by pixel
     11  SD
     12  CoV
     13  Min
     14  Max
     15  90th percentile
     16  95th percentile

    Stats (rows 10–16) are formulas referencing '3. Payout Amounts' (rows 10..N). Freeze at F9, after metadata.
    """
    cols = _resolve_cols(df)
    if not cols["pixel_key"] or not cols["year"]:
        raise ValueError("Dataframe must include Pixel_ID and Year.")

    pixel_col = cols["pixel_key"]; year_col = cols["year"]
    pixel_order = sorted(df[pixel_col].dropna().unique().tolist())
    year_list   = sorted(df[year_col].dropna().unique().tolist())

    # For area coloring we read Area from df (value only; display still mirrors Sheet 3)
    area_by_pixel: Dict[object, Optional[str]] = {}
    if cols["area"]:
        for pix in pixel_order:
            area_by_pixel[pix] = _first_non_null(df[df[pixel_col] == pix][cols["area"]])
    else:
        for pix in pixel_order:
            area_by_pixel[pix] = None

    wb = wb or Workbook()
    ws = wb.active if (wb.active and ws_title_is_default(wb.active.title)) else wb.create_sheet()
    ws.title = sheet_name

    bold = Font(bold=True)
    left = Alignment(horizontal="left")

    COL_LABELS = 5   # E
    COL_FIRST  = 6   # F

    # Freeze at F9 as requested
    ws.freeze_panes = ws.cell(row=9, column=COL_FIRST)

    # Label column E (fixed, non-dynamic)
    labels = [
        "Pixel count",
        "Loan amounts (USD)",
        "Sum insured",
        "Area",
        "Region",
        "Pixel Lon",
        "Pixel Lat",
        "Pixel ID",
        "Total",
        "Average Payout by pixel",
        "SD",
        "CoV",
        "Min",
        "Max",
        "90th percentile",
        "95th percentile",
    ]
    for i, lab in enumerate(labels, start=1):
        cell = ws.cell(row=i, column=COL_LABELS)
        cell.value = lab
        cell.font = bold
        cell.alignment = left

    sheet3 = "3. Payout Amounts"
    row_first_data = 10
    row_end = row_first_data + len(year_list) - 1

    last_col = COL_FIRST + len(pixel_order) - 1

    # Area fill palette (same as Sheet 3)
    area_colors_hex = {
        "Northern Zone": "1F77B4",
        "Central Zone": "2CA02C",
        "Lake Zone": "FF7F0E",
        "Western Zone": "9467BD",
        "Southern Highlands Zone": "8C564B",
        "Coastal Zone": "17BECF",
        "Zanzibar (Islands)": "7F7F7F",
    }

    def _color_for_area(val: Optional[str]) -> Optional[str]:
        if not val:
            return None
        # exact match first
        hx = area_colors_hex.get(str(val))
        if hx:
            return hx
        # case-insensitive fallback
        sval = str(val).strip().lower()
        for k, v in area_colors_hex.items():
            if sval == k.lower():
                return v
        return None

    # Build columns F→ (per pixel)
    for j, pix in enumerate(pixel_order):
        c = COL_FIRST + j
        colL = get_column_letter(c)

        # 1 Pixel count (values 1..N)
        ws.cell(row=1, column=c).value = j + 1
        ws.cell(row=1, column=c).number_format = "0"

        # 2 Loan amounts (USD) — mirror Sheet 3 row 3
        ws.cell(row=2, column=c).value = f"='{sheet3}'!{colL}3"
        ws.cell(row=2, column=c).number_format = "#,##0"

        # 3 Sum insured — mirror Sheet 3 row 4
        ws.cell(row=3, column=c).value = f"='{sheet3}'!{colL}4"
        ws.cell(row=3, column=c).number_format = "#,##0"

        # 4 Area — mirror Sheet 3 row 5 + fill color
        area_cell = ws.cell(row=4, column=c)
        area_cell.value = f"='{sheet3}'!{colL}5"
        hx = _color_for_area(area_by_pixel.get(pix))
        if hx:
            area_cell.fill = _fill(hx)

        # 5 Region — mirror Sheet 3 row 6
        ws.cell(row=5, column=c).value = f"='{sheet3}'!{colL}6"

        # 6 Pixel Lon — mirror Sheet 3 row 7
        ws.cell(row=6, column=c).value = f"='{sheet3}'!{colL}7"

        # 7 Pixel Lat — mirror Sheet 3 row 8
        ws.cell(row=7, column=c).value = f"='{sheet3}'!{colL}8"

        # 8 Pixel ID — mirror Sheet 3 header row 9
        ws.cell(row=8, column=c).value = f"='{sheet3}'!{colL}9"

        # 9 Total — intentionally blank per request (no formulas/values)
        ws.cell(row=9, column=c).value = ""

        # Stats 10–16 — formulas based on Sheet 3 payouts grid
        col_rng_usd = f"'{sheet3}'!{colL}{row_first_data}:{colL}{row_end}"

        # 10 Average Payout by pixel
        ws.cell(row=10, column=c).value = f"=IF(COUNT({col_rng_usd})=0,\"\",AVERAGE({col_rng_usd}))"
        ws.cell(row=10, column=c).number_format = "#,##0"

        # 11 SD (needs at least 2)
        ws.cell(row=11, column=c).value = f"=IF(COUNT({col_rng_usd})<=1,\"\",STDEV({col_rng_usd}))"
        ws.cell(row=11, column=c).number_format = "#,##0"

        # 12 CoV = SD / Average (blank-safe)
        avg_ref = f"{colL}10"
        sd_ref  = f"{colL}11"
        ws.cell(row=12, column=c).value = f"=IF(OR(ISBLANK({avg_ref}),{avg_ref}=0,ISBLANK({sd_ref})),\"\",{sd_ref}/{avg_ref})"
        ws.cell(row=12, column=c).number_format = "0.00%"

        # 13 Min
        ws.cell(row=13, column=c).value = f"=IF(COUNT({col_rng_usd})=0,\"\",MIN({col_rng_usd}))"
        ws.cell(row=13, column=c).number_format = "#,##0"

        # 14 Max
        ws.cell(row=14, column=c).value = f"=IF(COUNT({col_rng_usd})=0,\"\",MAX({col_rng_usd}))"
        ws.cell(row=14, column=c).number_format = "#,##0"

        # 15 90th percentile
        ws.cell(row=15, column=c).value = f"=IF(COUNT({col_rng_usd})=0,\"\",PERCENTILE({col_rng_usd},0.9))"
        ws.cell(row=15, column=c).number_format = "#,##0"

        # 16 95th percentile
        ws.cell(row=16, column=c).value = f"=IF(COUNT({col_rng_usd})=0,\"\",PERCENTILE({col_rng_usd},0.95))"
        ws.cell(row=16, column=c).number_format = "#,##0"

    # Autosize columns
    _autosize(ws, 1, max(COL_LABELS, last_col))

    # Force full recalc on open
    try:
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    except Exception:
        try:
            wb.calculation.fullCalcOnLoad = True
        except Exception:
            pass

    return wb
