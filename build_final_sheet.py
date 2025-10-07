import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import matplotlib.pyplot as plt
from MakeExogenousExcelInputDataframe import load_and_merge, build_regional_statistics
from builder_excel_sheet1_fmt2 import build_excel_sheet1
from builder_excel_sheet2_formulas_fmt2 import build_excel_sheet2
from builder_excel_sheet3_formulas_fmt2 import build_excel_sheet3
from builder_excel_sheet4_formulas_fmt2 import build_excel_sheet4
from builder_excel_sheet5_formulas_fmt2 import build_excel_sheet5
from builder_excel_sheet6_formulas_fmt2 import build_excel_sheet6
from builder_excel_sheet7_8_9_formulas import build_excel_sheet7_chartdata, build_excel_sheet8_area_payout_chart, augment_sheet7_with_percentage_block, build_excel_sheet9_area_payout_pct_chart
from builder_excel_sheet10_formulas import build_excel_sheet10_diversification_chart
from builder_summary_sheet import build_excel_sheet0_summary

def add_chart_sheet(wb: Workbook, chart_func, df_data, df_final=None, sheet_name: str = "Chart"):
    """Add a chart sheet to the workbook"""
    ws = wb.create_sheet(title=sheet_name)
    
    # Use a simple file path in the current directory
    chart_path = f"temp_chart_{len(wb.worksheets)}.png"
    
    try:
        # Create the chart and save to file
        if df_final is not None:
            fig = chart_func(df_data, df_final, chart_path)
        else:
            fig = chart_func(df_data, chart_path)
        
        # Close the figure to free resources
        plt.close(fig)
        
        # Add image to worksheet
        img = Image(chart_path)
        img.anchor = 'A2'
        ws.add_image(img)
        
        # Add title
        ws['A1'].value = sheet_name
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
    finally:
        # Clean up chart file
        if os.path.exists(chart_path):
            try:
                pass
                #os.remove(chart_path)
            except:
                pass  # Don't fail if cleanup doesn't work
    
    return wb

def build_final_report(out_path: str = "output/final_report.xlsx") -> str:
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    
    df_final = load_and_merge()
    df_regional, df_regional_fmt = build_regional_statistics(df_final)
    
    # Create single workbook and pass it to all builders
    wb = Workbook()
    
    # Each builder adds their sheet to the same workbook
    wb = build_excel_sheet1(df_final, wb=wb)
    wb = build_excel_sheet2(df_final, wb=wb)
    wb = build_excel_sheet3(df_final, wb=wb)
    wb = build_excel_sheet4(df_final, wb=wb)
    wb = build_excel_sheet5(df_regional, df_regional_fmt, wb=wb)
    wb = build_excel_sheet6(df_regional, df_regional_fmt, wb=wb)
    
    # Add chart sheets
    # 7a) Chart Data (for Area Payouts)
    wb = build_excel_sheet7_chartdata(df_regional, wb)
    # 7b) add percentage block for Sheet 9
    wb = augment_sheet7_with_percentage_block(wb, sheet7_name="7. Chart Data")

    # 8) Visualization referencing Sheet 7
    wb = build_excel_sheet8_area_payout_chart(wb)

    # 9) chart (100% stacked) from that block
    wb = build_excel_sheet9_area_payout_pct_chart(wb, sheet7_name="7. Chart Data", sheet9_name="9. Area Payout %")

    # 10) Diversification Benefit Chart
    wb = build_excel_sheet10_diversification_chart(wb)

    # 11) summary sheet 
    wb = build_excel_sheet0_summary(wb)
    # Remove the default empty sheet if it still exists
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb['Sheet'])
    

    wb.save(out_path)
    return out_path

if __name__ == "__main__":
    build_final_report()
    
    # Remove temp charts png files
    for file in os.listdir('.'):
        if file.startswith('temp_chart_') and file.endswith('.png'):
            try:
                os.remove(file)
            except:
                pass  # Don't fail if cleanup doesn't work