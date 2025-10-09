# builder_summary_sheet.py

from typing import List, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

HEADER = Font(b=True, size=14)
NAME_RED_BOLD = Font(b=True, color="FF0000")
BOLD = Font(b=True)
NOT_BOLD = Font(b=False)
TOPLEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
TOPLEFT = Alignment(horizontal="left", vertical="top")
RIGHT = Alignment(horizontal="right", vertical="top")

DATA_LINES: List[Tuple[str, str]] = [
    ("1. Modelled Yield",
     "Worksheet reporting historical modeled yields by year and pixel. "
     "Yield expressed in kg per hectare. Year x refers to the production cycle "
     "starting in year x and harvesting in year x+1."),
    ("2. Payouts %",
     "Pixel-level payout percentage by year based on attach/detach thresholds from the header. "
     "Blanks are preserved when yield is missing; includes per-year and per-pixel statistics."),
    ("3. Payout Amounts",
     "Pixel-level payout amounts (USD) computed as Payout % × Sum Insured (Sum Insured derived from Loan Amount). "
     "Includes per-year and per-pixel statistics; average payout reported per pixel per year."),
    ("4. Pixel-by-Pixel Stats",
     "Per-pixel summary statistics (Average, SD, CoV, Min/Max/90th/95th) alongside pixel metadata "
     "(area, region, lon/lat, pixel ID)."),
    ("5. Regional Totals (%)",
     "Regional totals and percentage shares by year, plus an Overall Total column for national aggregates. "
     "Years without data remain blank."),
    ("6. Regional Totals (Stats Only)",
     "Region-level KPIs without time series: Average non-zero/blank pixel CoV, Area CoV, and Overall Total "
     "(national CoV) used for portfolio comparisons."),
    ("7. Chart Data",
     "Normalized chart inputs: Year + per-area payout totals and per-area shares (%), "
     "used by the charts that follow."),
]

CHART_LINES: List[Tuple[str, str]] = [
    ("8. Area Payout Chart",
     "Stacked column chart of payout amounts (USD) by area over time. Uses totals from '7. Chart Data'."),
    ("9. Area Payout %",
     "100% stacked column chart of area shares by year. Uses the percentage block from '7. Chart Data'."),
    ("10. Diversification Benefit",
     "Clustered CoV bars (Avg pixel, Area, National) with markers-only benefit (% reduction vs. national) "
     "on a secondary right-side axis."),
]


def _get_last_pixel_col_letter(ws_modelled_yield) -> str:
    """Find last used pixel column from Sheet 1 (header on row 9, data starts at F)."""
    PIXEL_ID_ROW = 9
    first = 6
    last = first
    maxc = ws_modelled_yield.max_column
    for c in range(first, maxc + 1):
        if ws_modelled_yield.cell(row=PIXEL_ID_ROW, column=c).value not in (None, ""):
            last = c
    return get_column_letter(last)


def _extract_crop_from_experiment_name(experiment_name: Optional[str]) -> str:
    """
    Parse the crop name from an experiment name like:
      'OutputCalendarDays180_Maize_1982_2021_SPARSE'
    We split on '_' and take the second token (index 1).
    """
    if not experiment_name:
        return "Unknown"
    parts = str(experiment_name).split("_")
    return parts[1] if len(parts) > 1 and parts[1] else "Unknown"


def build_excel_sheet0_summary(
    wb: Workbook,
    sheet_name: str = "0. Summary",
    experiment_name: Optional[str] = None,  # NEW: to populate B1 "Crop: <name>"
) -> Workbook:
    """
    Builds '0. Summary' with:
      - Section: Data (A column titles + merged B:J descriptions)
      - B1 shows 'Crop: <name>' in blue size 14 (parsed from experiment_name)
      - Section: Charts
      - Section: Summary Statistics (Region | # of Pixels | # of Farmers | Sum Insured | Average Payout)
      - Final Section: Data sources (two fixed entries)

    Workbook references:
      - Sheet 1 ("1. Modelled Yield"): region row=5, farmer count row=6, pixel data from F..Last
      - Sheet 3 ("3. Payout Amounts"): Sum Insured row=4, and stats row labeled "Average Payout by pixel" in column E
    """
    # Replace existing then create at index 0
    if sheet_name in [ws_.title for ws_ in wb.worksheets]:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(title=sheet_name, index=0)

    # Column widths
    ws.column_dimensions["A"].width = 28
    for col in range(2, 11):  # B..J for merged descriptions
        ws.column_dimensions[get_column_letter(col)].width = 16

    row = 1

    # ----- Section: Data -----
    ws.cell(row=row, column=1, value="Data").font = HEADER
    ws.cell(row=row, column=1).alignment = TOPLEFT_WRAP

    # B1: Crop: <name> (blue, size 14)
    crop = _extract_crop_from_experiment_name(experiment_name)
    c_b1 = ws.cell(row=1, column=2, value=f"Crop: {crop}")
    c_b1.font = Font(color="0000FF", size=14, bold=False)
    c_b1.alignment = Alignment(horizontal="left", vertical="top")

    row += 2

    for title, desc in DATA_LINES:
        name_cell = ws.cell(row=row, column=1, value=title)
        name_cell.font = NAME_RED_BOLD
        name_cell.alignment = TOPLEFT_WRAP

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        desc_cell = ws.cell(row=row, column=2, value=desc)
        desc_cell.alignment = TOPLEFT_WRAP
        row += 1

    row += 1  # spacer

    # ----- Section: Charts -----
    ws.cell(row=row, column=1, value="Charts").font = HEADER
    ws.cell(row=row, column=1).alignment = TOPLEFT_WRAP
    row += 2
    for title, desc in CHART_LINES:
        name_cell = ws.cell(row=row, column=1, value=title)
        name_cell.font = NAME_RED_BOLD
        name_cell.alignment = TOPLEFT_WRAP

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        desc_cell = ws.cell(row=row, column=2, value=desc)
        desc_cell.alignment = TOPLEFT_WRAP
        row += 1

    row += 2  # spacer

    # ----- Section: Summary Statistics -----
    ws.cell(row=row, column=1, value="Summary Statistics").font = HEADER
    ws.cell(row=row, column=1).alignment = TOPLEFT
    row += 2

    # Headers
    headers = ["Region", "# of Pixels", "# of Farmers", "Sum Insured", "Average Payout"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = BOLD
        c.alignment = TOPLEFT if j == 1 else RIGHT
    ws.column_dimensions["E"].width = 18
    row += 1

    # Region list from Sheet 1 Region row
    if "1. Modelled Yield" not in [w.title for w in wb.worksheets]:
        regions: List[str] = []
        last_letter = "F"
    else:
        ws1 = wb["1. Modelled Yield"]
        last_letter = _get_last_pixel_col_letter(ws1)
        regions = []
        for c in range(6, ws1.max_column + 1):
            rv = ws1.cell(row=5, column=c).value  # Region row in Sheet 1
            if rv not in (None, ""):
                regions.append(str(rv))
        regions = sorted(list(set(regions)))

    # Absolute ranges used in formulas
    # Sheet 1 (Modelled Yield)
    REG_ROW = 5
    FCT_ROW = 6  # Farmer count
    reg_rng = f"'1. Modelled Yield'!$F${REG_ROW}:'1. Modelled Yield'!${last_letter}${REG_ROW}"
    fct_rng = f"'1. Modelled Yield'!$F${FCT_ROW}:'1. Modelled Yield'!${last_letter}${FCT_ROW}"

    # Sheet 3 (Payout Amounts)
    SUMINS_ROW_S3 = 4  # Sum Insured row
    sumins_rng = f"'3. Payout Amounts'!$F${SUMINS_ROW_S3}:'3. Payout Amounts'!${last_letter}${SUMINS_ROW_S3}"

    # Dynamic row for "Average Payout by pixel" on Sheet 3 (label in column E)
    avg_row_ref = "MATCH(\"Average Payout by pixel\",'3. Payout Amounts'!$E:$E,0)"
    avg_row_rng = f"INDEX('3. Payout Amounts'!$F:${last_letter},{avg_row_ref},0)"

    start_table_row = row

    # Body rows
    for rname in regions:
        # Region (not bold)
        a = ws.cell(row=row, column=1, value=rname)
        a.font = NOT_BOLD
        a.alignment = TOPLEFT

        # # of Pixels
        b = ws.cell(row=row, column=2, value=f"=SUMPRODUCT(--({reg_rng}=A{row}))")
        b.font = NOT_BOLD; b.alignment = RIGHT; b.number_format = "#,##0"

        # # of Farmers
        c = ws.cell(row=row, column=3, value=f"=SUMPRODUCT(--({reg_rng}=A{row}),{fct_rng})")
        c.font = NOT_BOLD; c.alignment = RIGHT; c.number_format = "#,##0"

        # Sum Insured (from Sheet 3 row 4)
        d = ws.cell(row=row, column=4, value=f"=SUMPRODUCT(--({reg_rng}=A{row}),{sumins_rng})")
        d.font = NOT_BOLD; d.alignment = RIGHT; d.number_format = "#,##0"

        # Average Payout (sum of per-pixel averages for region, from Sheet 3 stats row)
        e = ws.cell(row=row, column=5, value=f"=SUMPRODUCT(--({reg_rng}=A{row}),{avg_row_rng})")
        e.font = NOT_BOLD; e.alignment = RIGHT; e.number_format = "#,##0"

        row += 1

    end_table_row = row - 1

    row += 1  # blank line

    # Totals (labels bold; numbers not bold)
    ws.cell(row=row, column=1, value="Total # Pixels").font = BOLD
    t_pix = ws.cell(row=row, column=2, value=f"=SUM(B{start_table_row}:B{end_table_row})")
    t_pix.font = NOT_BOLD; t_pix.alignment = RIGHT; t_pix.number_format = "#,##0"
    row += 1

    ws.cell(row=row, column=1, value="Total # of Farmers").font = BOLD
    t_far = ws.cell(row=row, column=2, value=f"=SUM(C{start_table_row}:C{end_table_row})")
    t_far.font = NOT_BOLD; t_far.alignment = RIGHT; t_far.number_format = "#,##0"
    row += 1

    ws.cell(row=row, column=1, value="Total Sum Insured").font = BOLD
    t_sumins = ws.cell(row=row, column=2, value=f"=SUM(D{start_table_row}:D{end_table_row})")
    t_sumins.font = NOT_BOLD; t_sumins.alignment = RIGHT; t_sumins.number_format = "#,##0"
    row += 1

    # Expected Loss (% sum insured) from Sheet 3
    avg_payout_sum = f"SUM(INDEX('3. Payout Amounts'!$F:${last_letter},{avg_row_ref},0))"
    suminsured_sum = f"SUM('3. Payout Amounts'!$F${SUMINS_ROW_S3}:'3. Payout Amounts'!${last_letter}${SUMINS_ROW_S3})"
    ws.cell(row=row, column=1, value="Expected Loss (% sum insured)").font = BOLD
    el = ws.cell(row=row, column=2, value=f"=IFERROR({avg_payout_sum}/{suminsured_sum},\"\")")
    el.font = NOT_BOLD; el.alignment = RIGHT; el.number_format = "0.0%"
    row += 2  # spacer before data sources

    # ----- Final Section: Data sources -----
    ws.cell(row=row, column=1, value="Data sources").font = HEADER
    ws.cell(row=row, column=1).alignment = TOPLEFT
    row += 2

    ws.cell(row=row, column=1, value="1. Precipitation: CHIRPS").alignment = TOPLEFT
    row += 1
    ws.cell(row=row, column=1, value="2. Temperature: ERA5_Land").alignment = TOPLEFT

    # Freeze panes
    ws.freeze_panes = "A4"

    return wb
