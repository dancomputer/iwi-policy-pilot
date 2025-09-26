import math
from typing import Optional, Dict, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def _norm(s: str) -> str:
    return str(s).strip().lower().replace(" ", "").replace("_", "")

def _resolve_cols(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    norm_to_orig = {_norm(c): c for c in df.columns}
    def pick(*aliases: str) -> Optional[str]:
        for a in aliases:
            k = _norm(a)
            if k in norm_to_orig:
                return norm_to_orig[k]
        return None
    return {
        "pixel_key": pick("Pixel_ID", "pixelid", "pixel"),
        "year": pick("year"),
        "payout_amount": pick("PayoutAmountBase", "payoutamountbase", "payout_amount", "payoutamount"),
        "payout_percent": pick("PayoutsPercent", "payoutspercent", "percent_payout"),
        "loan_amount": pick("Loan_Amount", "loanamount"),
        "sum_insured": pick("Sum_Insured", "suminsured"),
        "area": pick("area"),
        "region": pick("region"),
        "pixel_lon": pick("lon", "longitude"),
        "pixel_lat": pick("lat", "latitude"),
        "pixel_id": pick("Pixel_ID", "pixelid"),
    }

def _first_non_null(s: pd.Series):
    s2 = s.dropna()
    return s2.iloc[0] if not s2.empty else None

def autosize_columns(ws, col_start: int, col_end: int, min_width: int = 8, max_width: int = 45):
    for col in range(col_start, col_end + 1):
        m = 0
        for vals in ws.iter_cols(min_col=col, max_col=col, min_row=1, max_row=ws.max_row, values_only=True):
            for v in vals:
                if v is None:
                    continue
                m = max(m, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, min(max_width, m + 2))

def ws_title_is_default(title: str) -> bool:
    return str(title).lower().startswith("sheet")

def build_excel_sheet3(
    df: pd.DataFrame,
    wb: Optional[Workbook] = None,
    sheet_name: str = "3. Payout Amounts (USD)"
) -> Workbook:
    cols = _resolve_cols(df)

    # If payout amount missing but percent + sum insured exist, derive it.
    if not cols["payout_amount"] and cols["payout_percent"] and cols["sum_insured"]:
        df = df.copy()
        df["_computed_payout_amount"] = df[cols["payout_percent"]] * df[cols["sum_insured"]]
        cols["payout_amount"] = "_computed_payout_amount"

    required = {k: cols[k] for k in ("pixel_key", "year", "payout_amount", "loan_amount", "sum_insured")}
    missing = [k for k, v in required.items() if not v]
    if missing:
        raise ValueError(f"Missing required columns: {missing}. Available: {list(df.columns)}")

    pixel_col = cols["pixel_key"]
    year_col = cols["year"]
    payout_col = cols["payout_amount"]

    pivot = df.pivot_table(index=year_col, columns=pixel_col, values=payout_col, aggfunc="first").sort_index()
    pixel_order: List = [c for c in pivot.columns if not pd.isna(c)]

    # Metadata per pixel
    meta: Dict = {}
    for pix in pixel_order:
        sub = df[df[pixel_col] == pix]
        meta[pix] = {
            "loan": _first_non_null(sub[cols["loan_amount"]]) if cols["loan_amount"] else None,
            "sum_insured": _first_non_null(sub[cols["sum_insured"]]) if cols["sum_insured"] else None,
            "area": _first_non_null(sub[cols["area"]]) if cols["area"] else None,
            "region": _first_non_null(sub[cols["region"]]) if cols["region"] else None,
            "lon": _first_non_null(sub[cols["pixel_lon"]]) if cols["pixel_lon"] else None,
            "lat": _first_non_null(sub[cols["pixel_lat"]]) if cols["pixel_lat"] else None,
            "pixelid": _first_non_null(sub[cols["pixel_id"]]) if cols["pixel_id"] else pix,
        }

    # Per-year statistics
    year_avg = pivot.mean(axis=1, skipna=True)
    year_sd = pivot.std(axis=1, ddof=0, skipna=True)

    # By-pixel summary
    by_pixel_avg = pivot.mean(axis=0, skipna=True)
    by_pixel_sd = pivot.std(axis=0, ddof=0, skipna=True)
    by_pixel_min = pivot.min(axis=0, skipna=True)
    by_pixel_max = pivot.max(axis=0, skipna=True)
    by_pixel_p90 = pivot.quantile(0.90, axis=0, interpolation="linear")
    by_pixel_p95 = pivot.quantile(0.95, axis=0, interpolation="linear")

    avg_sd = float(year_sd.mean(skipna=True)) if len(year_sd) else None
    avg_payout_overall = float(year_avg.mean(skipna=True)) if len(year_avg) else None
    overall_sd = float(year_avg.std(ddof=0)) if len(year_avg) else None

    total_loan = sum([meta[p]["loan"] for p in pixel_order if meta[p]["loan"] is not None])
    total_pixels = len(pixel_order)

    wb = wb or Workbook()
    if wb.active and ws_title_is_default(wb.active.title) and wb.active.max_row == 1:
        ws = wb.active
    else:
        ws = wb.create_sheet()
    ws.title = sheet_name

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    left = Alignment(horizontal="left")

    # Freeze top 10 rows + columns A–E (pixel metadata + labels)
    # First scrollable cell should be F11 (row 11 is first data row after removing old header row)
    ws.freeze_panes = ws.cell(row=11, column=6)  # F11

    # Color palette
    area_colors_hex = {
        "Northern Zone": "1F77B4",
        "Central Zone": "2CA02C",
        "Lake Zone": "FF7F0E",
        "Western Zone": "9467BD",
        "Southern Highlands Zone": "8C564B",
        "Coastal Zone": "17BECF",
        "Zanzibar (Islands)": "7F7F7F",
    }
    def to_fill(hex6: str) -> PatternFill:
        return PatternFill(fill_type="solid", start_color=f"FF{hex6}", end_color=f"FF{hex6}")

    # Layout
    col_label = 5      # E (Year values - no header text)
    first_data_col = 6 # F
    row_title = 2
    row_meta_start = 3
    # We REMOVE the former header row (row_meta_start + 8). Place SD / Average on the Pixel ID metadata row.
    row_pixel_id = row_meta_start + 7  # This is row 10
    row_header = row_pixel_id          # Reuse row 10 for SD / Average
    row_data_start = row_header + 1    # Data now begins at row 11

    # Note merged D1:D3
    note = (
        "Note: Loan amounts are given as average loan amount across pixels within a given region, "
        "as pixel-level loan amount info not available yet."
    )
    ws.merge_cells(start_row=1, start_column=4, end_row=3, end_column=4)  # D1:D3
    nc = ws.cell(row=1, column=4)
    nc.value = note
    nc.alignment = Alignment(wrap_text=True, vertical="top")

    # Title
    title_cell = ws.cell(row=row_title, column=col_label)
    title_cell.value = "PAYOUT AMOUNTS (USD)"
    title_cell.font = bold
    title_cell.fill = PatternFill(fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00")

    # Metadata rows
    # Pixel count
    ws.cell(row=row_meta_start + 0, column=col_label).value = "Pixel count"
    for j, _ in enumerate(pixel_order):
        c = ws.cell(row=row_meta_start + 0, column=first_data_col + j)
        c.value = j + 1
        c.alignment = center

    # Total loan amounts (aggregate + per pixel) with currency
    ws.cell(row=row_meta_start + 1, column=1).value = "Total loan amounts"
    agg_loan_cell = ws.cell(row=row_meta_start + 1, column=3)
    agg_loan_cell.value = total_loan
    agg_loan_cell.number_format = "$#,##0"
    ws.cell(row=row_meta_start + 1, column=col_label).value = "Loan amounts (USD)"
    for j, pix in enumerate(pixel_order):
        v = meta[pix]["loan"]
        c = ws.cell(row=row_meta_start + 1, column=first_data_col + j)
        c.value = v
        c.number_format = "$#,##0"

    # Sum insured
    ws.cell(row=row_meta_start + 2, column=col_label).value = "Sum insured"
    for j, pix in enumerate(pixel_order):
        v = meta[pix]["sum_insured"]
        c = ws.cell(row=row_meta_start + 2, column=first_data_col + j)
        c.value = v
        c.number_format = "#,##0"

    # Total pixels + Area
    ws.cell(row=row_meta_start + 3, column=1).value = "Total number of pixels"
    ws.cell(row=row_meta_start + 3, column=3).value = total_pixels
    ws.cell(row=row_meta_start + 3, column=col_label).value = "Area"
    for j, pix in enumerate(pixel_order):
        cell = ws.cell(row=row_meta_start + 3, column=first_data_col + j)
        v = meta[pix]["area"] or ""
        cell.value = v
        hx = None
        for k, hexv in area_colors_hex.items():
            if str(v).lower() == k.lower():
                hx = hexv
                break
        if hx:
            cell.fill = to_fill(hx)

    # Region
    ws.cell(row=row_meta_start + 4, column=col_label).value = "Region"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=row_meta_start + 4, column=first_data_col + j).value = meta[pix]["region"] or ""

    # Pixel Lon
    ws.cell(row=row_meta_start + 5, column=col_label).value = "Pixel Lon"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=row_meta_start + 5, column=first_data_col + j).value = meta[pix]["lon"]

    # Pixel Lat
    ws.cell(row=row_meta_start + 6, column=col_label).value = "Pixel Lat"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=row_meta_start + 6, column=first_data_col + j).value = meta[pix]["lat"]

    # Pixel ID row (row 10)
    ws.cell(row=row_meta_start + 7, column=col_label).value = "Pixel ID"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=row_meta_start + 7, column=first_data_col + j).value = meta[pix]["pixelid"]

    # Place SD / Average labels on the same row (row 10) instead of a new header row
    ws.cell(row=row_header, column=2).value = "SD"
    ws.cell(row=row_header, column=4).value = "Average"
    ws.cell(row=row_header, column=2).alignment = center
    ws.cell(row=row_header, column=4).alignment = center

    # Data rows (start now at row 11)
    years = list(pivot.index)
    for i, y in enumerate(years, start=1):
        r = row_data_start + i - 1
        # SD
        c_sd = ws.cell(row=r, column=2)
        v_sd = year_sd.iat[i - 1] if len(year_sd) >= i else None
        c_sd.value = float(v_sd) if pd.notnull(v_sd) else None
        c_sd.number_format = "#,##0"
        c_sd.font = bold
        c_sd.alignment = center
        # Average
        c_avg = ws.cell(row=r, column=4)
        v_avg = year_avg.iat[i - 1] if len(year_avg) >= i else None
        c_avg.value = float(v_avg) if pd.notnull(v_avg) else None
        c_avg.number_format = "#,##0"
        c_avg.font = bold
        c_avg.alignment = center
        # Year (no header label)
        try:
            ws.cell(row=r, column=col_label).value = int(y)
        except Exception:
            ws.cell(row=r, column=col_label).value = y
        # Pixel payouts
        for j in range(len(pixel_order)):
            pv = pivot.iat[i - 1, j]
            c = ws.cell(row=r, column=first_data_col + j)
            c.value = float(pv) if pd.notnull(pv) else None
            c.number_format = "#,##0"

    # Summary rows
    r_sum1 = row_data_start + len(years)
    ws.cell(row=r_sum1, column=1).value = "Average SD"
    c_avg_sd = ws.cell(row=r_sum1, column=2)
    c_avg_sd.value = avg_sd if avg_sd is not None else None
    c_avg_sd.number_format = "#,##0"
    c_avg_sd.alignment = center

    ws.cell(row=r_sum1, column=3).value = "Average payout (per pixel per year)"
    c_avg_overall = ws.cell(row=r_sum1, column=4)
    c_avg_overall.value = avg_payout_overall if avg_payout_overall is not None else None
    c_avg_overall.number_format = "#,##0"
    c_avg_overall.alignment = center

    ws.cell(row=r_sum1, column=col_label).value = "Average Payout per pixel"
    for j in range(len(pixel_order)):
        v = by_pixel_avg.iloc[j]
        c = ws.cell(row=r_sum1, column=first_data_col + j)
        c.value = float(v) if pd.notnull(v) else None
        c.number_format = "#,##0"

    r_sd = r_sum1 + 1
    ws.cell(row=r_sd, column=1).value = "Overall SD"
    c_overall_sd = ws.cell(row=r_sd, column=2)
    c_overall_sd.value = overall_sd if overall_sd is not None else None
    c_overall_sd.number_format = "#,##0"
    c_overall_sd.alignment = center

    ws.cell(row=r_sd, column=col_label).value = "SD"
    for j in range(len(pixel_order)):
        v = by_pixel_sd.iloc[j]
        c = ws.cell(row=r_sd, column=first_data_col + j)
        c.value = float(v) if pd.notnull(v) else None
        c.number_format = "#,##0"

    for offset, label, series in [
        (1, "Min", by_pixel_min),
        (2, "Max", by_pixel_max),
        (3, "90th percentile", by_pixel_p90),
        (4, "95th percentile", by_pixel_p95),
    ]:
        r = r_sd + offset
        ws.cell(row=r, column=col_label).value = label
        for j in range(len(pixel_order)):
            v = series.iloc[j]
            c = ws.cell(row=r, column=first_data_col + j)
            c.value = float(v) if pd.notnull(v) else None
            c.number_format = "#,##0"

    # Bold label formatting
    bold_rows = list(range(row_meta_start, row_meta_start + 8)) + [
        row_header, r_sum1, r_sd, r_sd + 1, r_sd + 2, r_sd + 3, r_sd + 4
    ]
    for rr in bold_rows:
        cellE = ws.cell(row=rr, column=col_label)
        if cellE.value:
            cellE.font = bold
            cellE.alignment = left
    for rr in (row_meta_start + 1, row_meta_start + 3, r_sum1, r_sd):
        for cc in (1, 3):
            c = ws.cell(row=rr, column=cc)
            if c.value is not None:
                c.font = bold

    # Autosize
    autosize_columns(ws, 1, max(first_data_col + len(pixel_order) - 1, col_label))
    return wb

# Example usage:
# wb = build_payout_amounts_sheet(df_final)
# wb.save("policy_pilot_payout_amounts.xlsx")