import math
from typing import Optional, Dict, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

def _norm(s: str) -> str:
    return str(s).strip().lower().replace(" ", "").replace("_", "")

def _resolve_cols(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    norm_to_orig = {_norm(c): c for c in df.columns}
    def pick(*aliases: str) -> Optional[str]:
        for a in aliases:
            k = _norm(a)
            if k in norm_to_orig:
                return norm_to_orig[k]
        return None
    return {
        "pixel_key": pick("Pixel_ID", "pixelid", "pixel"),
        "year": pick("Year", "year"),
        "attach": pick("Attach", "attach_threshold", "attach_kg_ha"),
        "detach": pick("Detach", "detach_threshold", "detach_kg_ha"),
        "area": pick("Area", "area_ha", "hectares"),
        "region": pick("Region"),
        "pixel_lon": pick("lon", "longitude"),
        "pixel_lat": pick("lat", "latitude"),
        "pixel_id": pick("Pixel_ID", "pixelid"),
    }

def _first_non_null(s: pd.Series):
    s2 = s.dropna()
    return s2.iloc[0] if not s2.empty else None

def to_fill(hex6: str) -> PatternFill:
    return PatternFill(fill_type="solid", start_color=f"FF{hex6}", end_color=f"FF{hex6}")

def autosize_columns(ws, col_start: int, col_end: int, min_width: int = 8, max_width: int = 40):
    for col in range(col_start, col_end + 1):
        m = 0
        for vals in ws.iter_cols(min_col=col, max_col=col, min_row=1, max_row=ws.max_row, values_only=True):
            for v in vals:
                if v is None:
                    continue
                m = max(m, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, min(max_width, m + 2))

def ws_title_is_default(title: str) -> bool:
    return str(title).lower().startswith("sheet")

def build_excel_sheet2(df: pd.DataFrame, wb: Optional[Workbook] = None, sheet_name: str = "2. Payouts %") -> Workbook:
    """
    Sheet 2: Payouts Percent
    - Starts at row 1 (title), metadata rows 2..8, header 9, data 10 (aligned with '1. Modelled Yield').
    - Metadata rows (Pixel count, Attach, Detach, Area, Region, Lon, Lat) are VALUES.
    - ALL other cells (grid + stats) are FORMULAS (blank-safe).
    - Grid formulas reference '1. Modelled Yield' (same row/col) + Attach/Detach from this sheet.
    """
    cols = _resolve_cols(df)
    if not cols["pixel_key"] or not cols["year"]:
        raise ValueError("Dataframe must include Pixel_ID and Year.")

    pixel_col = cols["pixel_key"]
    year_col = cols["year"]

    # Order pixels and years
    pixel_order = sorted(df[pixel_col].dropna().unique().tolist())
    year_list = sorted(df[year_col].dropna().unique().tolist())

    # Per-pixel metadata (values)
    meta: Dict[object, Dict[str, Optional[object]]] = {}
    for pix in pixel_order:
        sub = df[df[pixel_col] == pix]
        meta[pix] = {
            "attach": _first_non_null(sub[cols["attach"]]) if cols["attach"] else None,
            "detach": _first_non_null(sub[cols["detach"]]) if cols["detach"] else None,
            "area":   _first_non_null(sub[cols["area"]])   if cols["area"]   else None,
            "region": _first_non_null(sub[cols["region"]]) if cols["region"] else None,
            "lon":    _first_non_null(sub[cols["pixel_lon"]]) if cols["pixel_lon"] else None,
            "lat":    _first_non_null(sub[cols["pixel_lat"]]) if cols["pixel_lat"] else None,
            "pixelid":_first_non_null(sub[cols["pixel_id"]]) if cols["pixel_id"] else pix,
        }

    wb = wb or Workbook()
    ws = wb.active if (wb.active and wb.active.max_row == 1 and ws_title_is_default(wb.active.title)) else wb.create_sheet()
    ws.title = sheet_name

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    left = Alignment(horizontal="left")

    # ===== Layout (now starting at row 1) =====
    COL_YEAR_LABEL = 5   # E
    COL_FIRST_PIXEL = 6  # F
    ROW_TITLE       = 1
    ROW_META_START  = 2  # rows 2..8 for metadata
    ROW_PIXEL_ID    = 9  # header row
    ROW_FIRST_DATA  = 10 # data row (match Sheet 1)
    ws.freeze_panes = ws.cell(row=ROW_FIRST_DATA, column=COL_FIRST_PIXEL)

    # Title (row 1)
    ws.cell(row=ROW_TITLE, column=COL_YEAR_LABEL).value = "PAYOUTS % (fraction of sum insured)"
    ws.cell(row=ROW_TITLE, column=COL_YEAR_LABEL).font = bold
    ws.cell(row=ROW_TITLE, column=COL_YEAR_LABEL).alignment = left
    ws.cell(row=ROW_TITLE, column=COL_YEAR_LABEL).fill = to_fill("FFFF00")

    # Optional colors for Area
    area_colors_hex = {
        "Northern Zone": "1F77B4",
        "Central Zone": "2CA02C",
        "Lake Zone": "FF7F0E",
        "Western Zone": "9467BD",
        "Southern Highlands Zone": "8C564B",
        "Coastal Zone": "17BECF",
        "Zanzibar (Islands)": "7F7F7F",
    }

    # ===== Metadata labels (E), values in F→ (rows 2..8) =====
    ws.cell(row=ROW_META_START + 0, column=COL_YEAR_LABEL).value = "Pixel count"
    ws.cell(row=ROW_META_START + 1, column=COL_YEAR_LABEL).value = "Attach (kg per ha)"
    ws.cell(row=ROW_META_START + 2, column=COL_YEAR_LABEL).value = "Detach (kg per ha)"
    ws.cell(row=ROW_META_START + 3, column=COL_YEAR_LABEL).value = "Area"
    ws.cell(row=ROW_META_START + 4, column=COL_YEAR_LABEL).value = "Region"
    ws.cell(row=ROW_META_START + 5, column=COL_YEAR_LABEL).value = "Pixel Lon"
    ws.cell(row=ROW_META_START + 6, column=COL_YEAR_LABEL).value = "Pixel Lat"

    for j, pix in enumerate(pixel_order):
        c = COL_FIRST_PIXEL + j
        ws.cell(row=ROW_META_START + 0, column=c).value = j + 1
        ws.cell(row=ROW_META_START + 1, column=c).value = meta[pix]["attach"]
        ws.cell(row=ROW_META_START + 2, column=c).value = meta[pix]["detach"]

        # Area (with fill if matched)
        area_cell = ws.cell(row=ROW_META_START + 3, column=c)
        v_area = meta[pix]["area"] or ""
        area_cell.value = v_area
        hx = None
        if v_area:
            hx = area_colors_hex.get(str(v_area), None)
            if not hx:
                for k, val in area_colors_hex.items():
                    if str(v_area).strip().lower() == k.lower():
                        hx = val; break
        if hx:
            area_cell.fill = to_fill(hx)

        ws.cell(row=ROW_META_START + 4, column=c).value = meta[pix]["region"] or ""
        ws.cell(row=ROW_META_START + 5, column=c).value = meta[pix]["lon"]
        ws.cell(row=ROW_META_START + 6, column=c).value = meta[pix]["lat"]

    # ===== Header row (row 9) =====
    ws.cell(row=ROW_PIXEL_ID, column=2).value = "SD"
    ws.cell(row=ROW_PIXEL_ID, column=4).value = "Average"
    ws.cell(row=ROW_PIXEL_ID, column=COL_YEAR_LABEL).value = "Year"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=ROW_PIXEL_ID, column=COL_FIRST_PIXEL + j).value = meta[pix]["pixelid"] or str(pix)

    # Year labels in E (row 10..)
    for i, y in enumerate(year_list):
        r = ROW_FIRST_DATA + i
        try:
            ws.cell(row=r, column=COL_YEAR_LABEL).value = int(y)
        except Exception:
            ws.cell(row=r, column=COL_YEAR_LABEL).value = y

    # ===== GRID: payout% from Sheet 1 + Attach/Detach (blank-safe) =====
    sheet1_name = "1. Modelled Yield"
    last_col = COL_FIRST_PIXEL + len(pixel_order) - 1

    for i, _ in enumerate(year_list):
        r = ROW_FIRST_DATA + i

        # Per-year stats in A–D (blank-safe)
        row_rng = f"{get_column_letter(COL_FIRST_PIXEL)}{r}:{get_column_letter(last_col)}{r}"
        ws.cell(row=r, column=2).value = f"=IF(COUNT({row_rng})<=1,\"\",STDEV({row_rng}))"
        ws.cell(row=r, column=2).number_format = "0.00%"
        ws.cell(row=r, column=4).value = f"=IF(COUNT({row_rng})=0,\"\",AVERAGE({row_rng}))"
        ws.cell(row=r, column=4).number_format = "0.00%"

        # Grid cells
        for j, _ in enumerate(pixel_order):
            c = COL_FIRST_PIXEL + j
            colL = get_column_letter(c)
            y_ref = f"'{sheet1_name}'!{colL}{r}"          # same row/col as Sheet 1
            a_ref = f"{colL}{ROW_META_START + 1}"         # Attach at row 3
            d_ref = f"{colL}{ROW_META_START + 2}"         # Detach at row 4
            formula = (
                f"=IF(OR(ISBLANK({y_ref}),NOT(ISNUMBER({y_ref}))),\"\","
                f"IF(OR(ISBLANK({a_ref}),ISBLANK({d_ref})),\"\","
                f"IF(MAX(N({a_ref}),N({d_ref}))=MIN(N({a_ref}),N({d_ref})),NA(),"
                f"IF({y_ref}<=MIN(N({a_ref}),N({d_ref})),1,"
                f"IF({y_ref}>=MAX(N({a_ref}),N({d_ref})),0,"
                f"MAX(0,MIN(1,(MAX(N({a_ref}),N({d_ref}))-{y_ref})/"
                f"(MAX(N({a_ref}),N({d_ref}))-MIN(N({a_ref}),N({d_ref}))))))))))"
            )
            cell = ws.cell(row=r, column=c)
            cell.value = formula
            cell.number_format = "0.00%"

    # ===== Summary rows under the grid =====
    end_row = ROW_FIRST_DATA + len(year_list) - 1
    r_sum1 = end_row + 1

    ws.cell(row=r_sum1, column=1).value = "Average SD"
    ws.cell(row=r_sum1, column=1).font = bold
    ws.cell(row=r_sum1, column=2).value = f"=IF(COUNT(B{ROW_FIRST_DATA}:B{end_row})=0,\"\",AVERAGE(B{ROW_FIRST_DATA}:B{end_row}))"
    ws.cell(row=r_sum1, column=2).number_format = "0.00%"

    ws.cell(row=r_sum1, column=3).value = "Average payout (% of sum insured)"
    ws.cell(row=r_sum1, column=3).font = bold
    grid_rng = f"{get_column_letter(COL_FIRST_PIXEL)}{ROW_FIRST_DATA}:{get_column_letter(last_col)}{end_row}"
    ws.cell(row=r_sum1, column=4).value = f"=IF(COUNT({grid_rng})=0,\"\",AVERAGE({grid_rng}))"
    ws.cell(row=r_sum1, column=4).number_format = "0.00%"

    ws.cell(row=r_sum1, column=COL_YEAR_LABEL).value = "Average Payout by pixel"
    ws.cell(row=r_sum1, column=COL_YEAR_LABEL).font = bold
    for j in range(len(pixel_order)):
        c = COL_FIRST_PIXEL + j
        colL = get_column_letter(c)
        col_rng = f"{colL}{ROW_FIRST_DATA}:{colL}{end_row}"
        ws.cell(row=r_sum1, column=c).value = f"=IF(COUNT({col_rng})=0,\"\",AVERAGE({col_rng}))"
        ws.cell(row=r_sum1, column=c).number_format = "0.00%"

    r_sd = r_sum1 + 1
    ws.cell(row=r_sd, column=1).value = "Overall SD"
    ws.cell(row=r_sd, column=1).font = Font(bold=True)
    ws.cell(row=r_sd, column=2).value = f"=IF(COUNT(D{ROW_FIRST_DATA}:D{end_row})<=1,\"\",STDEV(D{ROW_FIRST_DATA}:D{end_row}))"
    ws.cell(row=r_sd, column=2).number_format = "0.00%"
    ws.cell(row=r_sd, column=COL_YEAR_LABEL).value = "SD"
    ws.cell(row=r_sd, column=COL_YEAR_LABEL).font = Font(bold=True)
    for j in range(len(pixel_order)):
        c = COL_FIRST_PIXEL + j
        colL = get_column_letter(c)
        col_rng = f"{colL}{ROW_FIRST_DATA}:{colL}{end_row}"
        ws.cell(row=r_sd, column=c).value = f"=IF(COUNT({col_rng})<=1,\"\",STDEV({col_rng}))"
        ws.cell(row=r_sd, column=c).number_format = "0.00%"

    for offset, label, fbuild in [
        (1, "Min", lambda rng: f"=IF(COUNT({rng})=0,\"\",MIN({rng}))"),
        (2, "Max", lambda rng: f"=IF(COUNT({rng})=0,\"\",MAX({rng}))"),
        (3, "90th percentile", lambda rng: f"=IF(COUNT({rng})=0,\"\",PERCENTILE({rng},0.9))"),
        (4, "95th percentile", lambda rng: f"=IF(COUNT({rng})=0,\"\",PERCENTILE({rng},0.95))"),
    ]:
        r = r_sd + offset
        ws.cell(row=r, column=COL_YEAR_LABEL).value = label
        ws.cell(row=r, column=COL_YEAR_LABEL).font = Font(bold=True)
        for j in range(len(pixel_order)):
            c = COL_FIRST_PIXEL + j
            colL = get_column_letter(c)
            col_rng = f"{colL}{ROW_FIRST_DATA}:{colL}{end_row}"
            ws.cell(row=r, column=c).value = fbuild(col_rng)
            ws.cell(row=r, column=c).number_format = "0.00%"

    # Styling
    for rr in range(ROW_META_START, ROW_PIXEL_ID + 1):
        ws.cell(row=rr, column=COL_YEAR_LABEL).font = bold
        ws.cell(row=rr, column=COL_YEAR_LABEL).alignment = left
    for cc in [2, 4]:
        ws.cell(row=ROW_PIXEL_ID, column=cc).font = bold
        ws.cell(row=ROW_PIXEL_ID, column=cc).alignment = center
    for j in range(len(pixel_order)):
        ws.cell(row=ROW_PIXEL_ID, column=COL_FIRST_PIXEL + j).font = bold
        ws.cell(row=ROW_PIXEL_ID, column=COL_FIRST_PIXEL + j).alignment = center

    last_col = max(COL_FIRST_PIXEL + len(pixel_order) - 1, COL_YEAR_LABEL)
    autosize_columns(ws, 1, last_col)

    # Force recalc on open (safe)
    try:
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    except Exception:
        try:
            wb.calculation.fullCalcOnLoad = True
        except Exception:
            pass

    return wb
