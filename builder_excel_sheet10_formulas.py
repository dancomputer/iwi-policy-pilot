from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.axis import ChartLines
from openpyxl.utils import get_column_letter
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties,RichTextProperties

def build_excel_sheet10_diversification_chart(
    wb,
    sheet6_name: str = "6. Regional Totals (Stats Only)",
    sheet10_name: str = "10. Diversification Benefit",
    sep: str = ","  # ";" for EU locales, "," for US
):
    if sheet6_name not in [ws.title for ws in wb.worksheets]:
        raise ValueError(f"{sheet6_name} must exist before creating Sheet 10.")
    ws6 = wb[sheet6_name]

    # --- Find CoV rows by label in column A ---
    avg_pix_cov_row = None   # "Average non-zero/blank pixel CoV"
    area_cov_row    = None   # "CoV"
    for r in range(1, ws6.max_row + 1):
        label = str(ws6.cell(row=r, column=1).value or "").strip().lower()
        if "average non-zero/blank pixel cov" in label:
            avg_pix_cov_row = r
        elif label == "cov":
            area_cov_row = r
    if avg_pix_cov_row is None or area_cov_row is None:
        raise ValueError("Sheet 6 must have 'Average non-zero/blank pixel CoV' and 'CoV' rows (labels in col A).")

    # --- Identify area total columns & Overall Total using rows 3 (Area) and 4 (Region) ---
    area_cols, area_names = [], []
    overall_col = None
    last_nonblank_col = None

    for j in range(2, ws6.max_column + 1):
        area_header = ws6.cell(row=3, column=j).value  # "Area"
        region_mark = ws6.cell(row=4, column=j).value  # "Region"
        if area_header not in (None, ""):
            last_nonblank_col = j
        name = str(area_header or "").strip()

        # Overall Total: row3 == "Overall Total" AND row4 blank
        if name.lower() == "overall total" and (region_mark in (None, "")):
            overall_col = j
            continue

        # Area total column: row4 equals "Total"
        if isinstance(region_mark, str) and region_mark.strip().lower() == "total":
            area_cols.append(j)
            area_names.append(name)

    # Fallback if not found: use the last nonblank header as "Overall Total"
    if overall_col is None:
        overall_col = last_nonblank_col

    if not area_cols or overall_col is None:
        # Build a minimal sheet with a note
        if sheet10_name in [ws.title for ws in wb.worksheets]:
            wb.remove(wb[sheet10_name])
        ws10 = wb.create_sheet(title=sheet10_name[:31])
        ws10["B3"] = "No area totals (row4='Total') or Overall Total column (row3='Overall Total') found on Sheet 6."
        return wb

    # --- Rebuild Sheet 10 helper table with ONLY those areas ---
    if sheet10_name in [ws.title for ws in wb.worksheets]:
        wb.remove(wb[sheet10_name])
    ws10 = wb.create_sheet(title=sheet10_name[:31])

    ws10["A1"] = "Area"
    ws10["B1"] = "Average CoV of pixels within Area"
    ws10["C1"] = "Area CoV"
    ws10["D1"] = "National CoV"
    ws10["E1"] = "Diversification Benefit"

    nat_cell_addr = f"'{sheet6_name}'!{get_column_letter(overall_col)}{area_cov_row}"

    for i, (name, jcol) in enumerate(zip(area_names, area_cols), start=2):
        col_letter = get_column_letter(jcol)
        ws10.cell(row=i, column=1, value=name)  # Area
        # B: Avg pixel CoV (Sheet 6, avg_pix_cov_row)
        ws10.cell(row=i, column=2).value = f"='{sheet6_name}'!{col_letter}{avg_pix_cov_row}"
        # C: Area CoV (Sheet 6, area_cov_row)
        ws10.cell(row=i, column=3).value = f"='{sheet6_name}'!{col_letter}{area_cov_row}"
        # D: National CoV (Overall Total column on CoV row)
        ws10.cell(row=i, column=4).value = f"={nat_cell_addr}"
        # E: Diversification Benefit = IF(C=0;""; 1 - D/C)
        ws10.cell(row=i, column=5).value = f"=IF(C{i}=0{sep}\"\"{sep}1-D{i}/B{i})"
        ws10.cell(row=i, column=5).number_format = "0%"

    last_row = 1 + len(area_cols)

    # --- Build chart: clustered columns + line on secondary % axis ---
    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.overlap = 0
    bar.gapWidth = 200
    bar.title = "Diversification Benefit : CoV Reduction due to Portfolio Effect"
    try:
        bar.title.tx.rich.p[0].pPr.defRPr.sz = 2000
        bar.title.tx.rich.p[0].pPr.defRPr.b = True
    except Exception:
        pass

    # Categories (areas) from helper table on Sheet 10
    cats = Reference(ws10, min_col=1, min_row=2, max_col=1, max_row=last_row)

    # --- Primary: clustered columns (Avg pixel CoV, Area CoV, National CoV) ---
    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.overlap = 0
    bar.gapWidth = 200
    bar.title = "Diversification Benefit : CoV Reduction due to Portfolio Effect"
    try:
        bar.title.tx.rich.p[0].pPr.defRPr.sz = 2000
        bar.title.tx.rich.p[0].pPr.defRPr.b = True
    except Exception:
        pass

    s_avg  = Reference(ws10, min_col=2, min_row=1, max_col=2, max_row=last_row)
    s_area = Reference(ws10, min_col=3, min_row=1, max_col=3, max_row=last_row)
    s_nat  = Reference(ws10, min_col=4, min_row=1, max_col=4, max_row=last_row)
    bar.add_data(s_avg,  titles_from_data=True)
    bar.add_data(s_area, titles_from_data=True)
    bar.add_data(s_nat,  titles_from_data=True)

    # Optional solid fills (safe RGB)
    for k, hex6 in enumerate(["5B9BD5", "FF7F0E", "A5A5A5"]):  # avg, area, national
        try:
            bar.series[k].graphicalProperties.solidFill = hex6
        except Exception:
            pass
    
    bar.legend.position = "b"
    bar.x_axis.majorTickMark = "out"
    bar.y_axis.majorTickMark = "out"
    bar.y_axis.title = "Coefficient of Variation"
    bar.y_axis.scaling.min = 0
    bar.set_categories(cats)

    # --- Secondary: benefit as markers-only on RIGHT axis (use crosses="max" trick) ---
    line = LineChart()
    line.y_axis.axId = 200  # distinct axis ID for secondary
    line.y_axis.title = "Reduction of Regional Average CoV due to Diversification Effects"
    line.y_axis.number_format = "0%"
    line.y_axis.scaling.min = 0
    line.y_axis.scaling.max = 1
    line.y_axis.majorGridlines = ChartLines()

    benefit = Reference(ws10, min_col=5, min_row=1, max_col=5, max_row=last_row)
    line.add_data(benefit, titles_from_data=True)
    line.set_categories(cats)

    # Markers only (no connecting line) + bigger dots
    try:
        s = line.series[0]
        s.marker = Marker(symbol="circle", size=20)
        s.graphicalProperties.line.noFill = True  # hide the line
        s.graphicalProperties.solidFill = "FFC000"  # gold markers
    except Exception:
        pass
    # --- Give the plot area margins so axis labels / legend don't collide ---
    # x,y,w,h are fractions of the chart frame. Tweak to taste.

    # <<< KEY LINE: put the secondary Y-axis on the RIGHT >>>
    line.y_axis.crosses = "max"   # primary Y crosses X at its maximum → secondary shows on the right

    bar.x_axis.delete = False
    bar.y_axis.delete = False
    line.y_axis.delete = False
   
    #line.y_axis.tickLblPos = "low"
    line.x_axis.tickLblPos = "low"
    #bar.y_axis.tickLblPos = "low"
    bar.x_axis.tickLblPos = "low"

    #Make text bigger
    LABEL_SZ = 1300  # 15pt
    LEGEND_SZ = 1500  # 15pt
    bar.legend.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=LEGEND_SZ)),
                     endParaRPr=CharacterProperties())],
        bodyPr=RichTextProperties()
    )


    #
    #manually position axis labels
    line.layout = Layout(
        manualLayout=ManualLayout(
            xMode="edge",
            yMode="edge",
            x=0.05,  # Move right for y-axis title clearance
            y=0.15,  # Lift plot area upward for x-axis title clearance
            w=0.9,
            h=0.8
        )
    )

    bar.layout = Layout(
        manualLayout=ManualLayout(
            xMode="edge",
            yMode="edge",
            x=0.05,  # Move right for y-axis title clearance
            y=0.15,  # Lift plot area upward for x-axis title clearance
            w=0.9,
            h=0.8
        )
    )

    line.y_axis.title.layout = Layout(
        manualLayout=ManualLayout(
            xMode="edge",
            yMode="edge",
            x=0.96  # Shift title to the right
        )
    )
    
    bar.y_axis.title.layout = Layout(
        manualLayout=ManualLayout(
            xMode="edge",
            yMode="edge",
            x=-0.15  # Shift title to the left
        )
    )

    cp = CharacterProperties(sz=LABEL_SZ)  # Where size goes from 100 till 40000
    bar.y_axis.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=LABEL_SZ)),
                     endParaRPr=CharacterProperties())],
        bodyPr=RichTextProperties()
    )
    bar.x_axis.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=LABEL_SZ)),
                     endParaRPr=CharacterProperties())],
        bodyPr=RichTextProperties()
    )
    line.y_axis.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=LABEL_SZ)),
                     endParaRPr=CharacterProperties())],
        bodyPr=RichTextProperties()
    )
    line.y_axis.title.tx.rich.p[0].pPr.defRPr.sz = 1400
    bar.y_axis.title.tx.rich.p[0].pPr.defRPr.sz = 1400
    #line.x_axis.delete = False
    bar += line
    bar.height = 18
    bar.width  = 32

 
    # --- Remove ALL gridlines ---
    bar.y_axis.majorGridlines = None
    bar.x_axis.majorGridlines = None
    line.y_axis.majorGridlines = None

    # --- Legend below, not overlaying plot ---
    bar.legend.position = "b"
    bar.legend.overlay = False



    ws10.add_chart(bar, "B3")
    return wb
