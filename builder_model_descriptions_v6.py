
from __future__ import annotations
import re
from pathlib import Path
from typing import Optional, List, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# === Layout to match Sheet 1 ===
col_label = 5          # E
first_data_col = 6     # F
row_title = 2
row_meta_start = 3

# === Area color palette (copied from Sheet 1) ===
AREA_COLORS_HEX = {
    "Northern Zone": "1F77B4",
    "Central Zone": "2CA02C",
    "Lake Zone": "FF7F0E",
    "Western Zone": "9467BD",
    "Southern Highlands Zone": "8C564B",
    "Coastal Zone": "17BECF",
    "Zanzibar (Islands)": "7F7F7F",
}

# --- Helpers ---
def _autosize(ws, start_col=1, end_col=None, min_width=8, max_width=40):
    if end_col is None:
        end_col = ws.max_column
    for col in range(start_col, end_col + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))

def _parse_segments(text: str):
    DESC_PATTERN = re.compile(
        r"\bDays\s*(?P<d0>-?\d+(?:\.\d+)?)\s*to\s*(?P<d1>-?\d+(?:\.\d+)?)\s*:\s*"
        r"(?P<var>[^:]+?)\s*from\s*(?P<v0>-?\d+(?:\.\d+)?)\s*to\s*(?P<v1>-?\d+(?:\.\d+)?)\s*"
        r"\[(?P<unit>[^\]]+)\]\s*--\s*Importance\s*of\s*(?P<imp>-?\d+(?:\.\d+)?)\s*%",
        re.IGNORECASE
    )
    s = text.strip()
    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1]
    s = re.sub(r"^\s*deliverable\s+model\s*:\s*", "", s, flags=re.IGNORECASE)
    parts = [p.strip() for p in s.split(";") if p.strip()]
    out = []
    for i, part in enumerate(parts, 1):
        m = DESC_PATTERN.search(part)
        if not m:
            out.append(dict(i=i, d0="", d1="", var=part, v0="", v1="", unit="", imp=""))
        else:
            out.append(dict(
                i=i,
                d0=m.group("d0"), d1=m.group("d1"),
                var=m.group("var").strip(),
                v0=m.group("v0"), v1=m.group("v1"),
                unit=m.group("unit").strip(),
                imp=m.group("imp"),
            ))
    return out

def _region_code(region: str) -> str:
    import pandas as pd
    if pd.isna(region):
        return "XX"
    s = str(region).strip().upper()
    return s[:2] if s else "XX"

TANZANIA_ZONES = {
    "Northern Zone": ["Arusha","Kilimanjaro","Manyara","Tanga"],
    "Central Zone": ["Dodoma","Singida","Tabora"],
    "Lake Zone": ["Geita","Kagera","Mara","Mwanza","Shinyanga","Simiyu"],
    "Western Zone": ["Kigoma","Katavi"],
    "Southern Highlands Zone": ["Iringa","Mbeya","Njombe","Rukwa","Ruvuma","Songwe"],
    "Coastal Zone": ["Dar es Salaam","Lindi","Morogoro","Mtwara","Pwani"],
    "Zanzibar (Islands)": ["Pemba North","Pemba South","Unguja North","Unguja South","Mjini Magharibi"],
}

def _area_from_region(region: str) -> str:
    region_low = str(region).lower()
    for area, regions in TANZANIA_ZONES.items():
        for reg in regions:
            rl = reg.lower()
            if region_low == rl or region_low in rl or rl in region_low:
                return area
    return "Unknown"

def _load_pixel_meta(csv_path: str) -> pd.DataFrame:
    df = pd.read_csv(csv_path)
    cols = {c.lower().strip(): c for c in df.columns}
    pixel = cols.get("pixel", "Pixel")
    lat = cols.get("latitude", "Latitude")
    lon = cols.get("longitude", "Longitude")
    region = cols.get("region", "Region")
    farmer = cols.get("farmer number", "Farmer Number")
    meta = pd.DataFrame({
        "pixel": pd.to_numeric(df[pixel], errors="coerce").astype("Int64"),
        "lat": df[lat],
        "lon": df[lon],
        "region": df[region],
        "farmer_count": df[farmer],
    })
    meta["area"] = meta["region"].apply(_area_from_region)
    meta["pixel_id_str"] = meta.apply(lambda r: f"{_region_code(r['region'])}{int(r['pixel'])}" if pd.notna(r['pixel']) else None, axis=1)
    return meta

def _kelvin_to_celsius(value):
    try:
        return float(value) - 273.15
    except Exception:
        return value

def _safe_float(x):
    try:
        return float(str(x).strip().replace('"','').replace("'",""))
    except Exception:
        # last resort: strip any non-numeric except exponent markers and dot
        import re
        s = re.sub(r"[^0-9eE+\-\.]", "", str(x))
        try:
            return float(s)
        except Exception:
            return None

def _build_r2_map(meta_dir: str, pixel_order: List[int]) -> Dict[int, float]:
    """Read per-pixel CSVs named pixel####.csv in meta_dir and extract the R2 column."""
    mp: Dict[int, float] = {}
    base = Path(meta_dir)
    for pix in pixel_order:
        f = base / f"pixel{pix:04d}.csv"
        if not f.exists():
            continue
        try:
            d = pd.read_csv(f)
        except Exception:
            # Try with engine python in case of strange quoting
            try:
                d = pd.read_csv(f, engine="python")
            except Exception:
                continue
        cols = {c.strip().lower(): c for c in d.columns}
        # Accept 'r2' or 'r^2' variants
        print("calculating for R2-Pred (change to R2 later!!!")
        r2col = cols.get("r2-pred") or cols.get("r^2") or cols.get("r_2")
        if not r2col:
            # Try fuzzy match
            for c in d.columns:
                if str(c).strip().lower().replace("^","").replace("_","") == "r2":
                    r2col = c
                    break
        if r2col:
            # take first non-null value
            series = d[r2col].dropna()
            if not series.empty:
                val = _safe_float(series.iloc[0])
                if val is not None:
                    mp[pix] = val
    return {pix: mp.get(pix) for pix in pixel_order}

def build_model_descriptions(
    df: pd.DataFrame,
    wb: Optional[Workbook] = None,
    sheet_name: str = "7. Model Descriptions",
    descriptions_dir: Optional[str] = None,
    pixel_meta_csv: str = "village_pixel_matches_maize-nkasi.csv",
    meta_dir: Optional[str] = None,
) -> Workbook:
    # Workbook setup
    if wb is None:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(title=sheet_name)

    left = Alignment(horizontal="left")
    center = Alignment(horizontal="center")
    bold = Font(bold=True)

    # Title
    title_cell = ws.cell(row=row_title, column=col_label, value="MODEL DESCRIPTIONS (parsed)")
    title_cell.font = bold
    title_cell.alignment = left
    title_cell.fill = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")

    # Pixel order and metadata strictly from CSV
    meta = _load_pixel_meta(pixel_meta_csv)
    pixel_order: List[int] = list(meta["pixel"].dropna().astype(int).tolist())

    # Pre-read segments per pixel
    base_dir = Path(descriptions_dir or r"C:\Users\danie\NecessaryM1InternshipCode\ProjectRice\OutputCalendarDays180_Maize_1982_2021_SPARSE\ThreeVariableContiguous-deliverable-model_plaintext-description")
    segs_by_pix: Dict[int, List[Dict[str, str]]] = {}
    for pix in pixel_order:
        fpath = base_dir / f"pixel{pix:04d}.txt"
        text = fpath.read_text(encoding="utf-8", errors="ignore") if fpath.exists() else ""
        segs_by_pix[pix] = _parse_segments(text)

    # R2 map using per-pixel CSVs
    meta_base = meta_dir or r"C:\Users\danie\NecessaryM1InternshipCode\ProjectRice\OutputCalendarDays180_Maize_1982_2021_SPARSE\ThreeVariableContiguous-deliverable-model_meta"
    r2_map = _build_r2_map(meta_base, pixel_order)

    # Metadata block labels
    labels = ["Pixel count","Area","Region","Farmer count","Pixel Lon","Pixel Lat","Pixel ID"]
    for idx, lab in enumerate(labels):
        r = row_meta_start + idx
        ws.cell(row=r, column=col_label, value=lab).font = bold
        ws.cell(row=r, column=col_label).alignment = left

    # Write column-wise values
    for j, pix in enumerate(pixel_order):
        col = first_data_col + j
        rowm = meta[meta["pixel"] == pix].iloc[0]

        # Pixel count = enumeration
        ws.cell(row=row_meta_start + 0, column=col, value=j + 1).alignment = center

        # Area + color
        area_name = rowm["area"] if pd.notna(rowm["area"]) else ""
        cell_area = ws.cell(row=row_meta_start + 1, column=col, value=area_name)
        cell_area.alignment = left
        if area_name:
            hex6 = None
            for k, v in AREA_COLORS_HEX.items():
                if str(area_name).strip().lower() == k.lower():
                    hex6 = v
                    break
            if hex6:
                cell_area.fill = PatternFill(fill_type="solid", start_color=hex6, end_color=hex6)

        # Region
        ws.cell(row=row_meta_start + 2, column=col, value=rowm["region"]).alignment = left
        # Farmer count
        ws.cell(row=row_meta_start + 3, column=col, value=rowm["farmer_count"]).alignment = center
        # Lon / Lat
        ws.cell(row=row_meta_start + 4, column=col, value=float(rowm["lon"]) if pd.notna(rowm["lon"]) else None).alignment = center
        ws.cell(row=row_meta_start + 5, column=col, value=float(rowm["lat"]) if pd.notna(rowm["lat"]) else None).alignment = center
        # Pixel ID
        ws.cell(row=row_meta_start + 6, column=col, value=rowm["pixel_id_str"]).alignment = center

        # Segment rows start immediately after metadata (no spacer)
        r0 = row_meta_start + 7
        segs = segs_by_pix[pix]
        for s in segs:
            unit = s['unit']
            is_kelvin = isinstance(unit, str) and unit.strip().lower() in ("k", "kelvin")
            if is_kelvin:
                unit_out = "Celsius"
                v0 = (float(s['v0']) - 0) if s['v0'] not in ("", None) else s['v0']
                v1 = (float(s['v1']) - 0) if s['v1'] not in ("", None) else s['v1']
            else:
                unit_out = unit
                v0, v1 = s['v0'], s['v1']

            labels_seg = [
                f"Day start (segment {s['i']})",
                f"Day end (segment {s['i']})",
                f"Variable (segment {s['i']})",
                f"Value from (segment {s['i']})",
                f"Value to (segment {s['i']})",
                f"Unit (segment {s['i']})",
                f"Importance % (segment {s['i']})",
            ]
            for k, lab in enumerate(labels_seg):
                rr = r0 + (s['i'] - 1) * 7 + k
                if j == 0:
                    ws.cell(row=rr, column=col_label, value=lab).font = bold
                    ws.cell(row=rr, column=col_label).alignment = left

            ws.cell(row=r0 + (s['i'] - 1) * 7 + 0, column=col, value=s['d0']).alignment = center
            ws.cell(row=r0 + (s['i'] - 1) * 7 + 1, column=col, value=s['d1']).alignment = center
            ws.cell(row=r0 + (s['i'] - 1) * 7 + 2, column=col, value=s['var']).alignment = left
            ws.cell(row=r0 + (s['i'] - 1) * 7 + 3, column=col, value=v0).alignment = center
            ws.cell(row=r0 + (s['i'] - 1) * 7 + 4, column=col, value=v1).alignment = center
            ws.cell(row=r0 + (s['i'] - 1) * 7 + 5, column=col, value=unit_out).alignment = center
            ws.cell(row=r0 + (s['i'] - 1) * 7 + 6, column=col, value=s['imp']).alignment = center

    # Final "R2" row
    r2_row_index = ws.max_row + 1
    ws.cell(row=r2_row_index, column=col_label, value="R2").font = Font(bold=True)
    ws.cell(row=r2_row_index, column=col_label).alignment = Alignment(horizontal="left")
    for j, pix in enumerate(pixel_order):
        col = first_data_col + j
        ws.cell(row=r2_row_index, column=col, value=r2_map.get(pix)).alignment = Alignment(horizontal="center")

    # Resizing & freeze
    for letter in ("A","B","C","D"):
        ws.column_dimensions[letter].width = 1.0
    ws.column_dimensions["E"].width = 18.0
    last_col = first_data_col + len(pixel_order) - 1
    for c in range(first_data_col, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 21.4
    ws.freeze_panes = "F10"

    _autosize(ws, 1, last_col)
    return wb
