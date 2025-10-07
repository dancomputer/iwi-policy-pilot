# %%
import math
from typing import Optional, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --- Area color palette (consistent across sheets) ---
AREA_COLORS_HEX = {
    "Northern Zone": "1F77B4",
    "Central Zone": "2CA02C",
    "Lake Zone": "FF7F0E",
    "Western Zone": "9467BD",
    "Southern Highlands Zone": "8C564B",
    "Coastal Zone": "17BECF",
    "Zanzibar (Islands)": "7F7F7F",
}

def to_fill(hex6: str) -> PatternFill:
    return PatternFill(fill_type="solid", start_color=f"FF{hex6}", end_color=f"FF{hex6}")

def _norm(s: str) -> str:
    return str(s).strip().lower().replace(" ", "").replace("_", "")

def _resolve_cols(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    norm_to_orig = {_norm(c): c for c in df.columns}
    def pick(*aliases: str) -> Optional[str]:
        for a in aliases:
            key = _norm(a)
            if key in norm_to_orig:
                return norm_to_orig[key]
        return None
    return {
        "pixel_key": pick("Pixel_ID", "pixel"),  # primary key
        "year": pick("year"),
        "yield": pick("Yield_Abs", "yield_abs", "yield"),
        "area": pick("area"),
        "region": pick("region"),
        "pixel_lon": pick("lon", "longitude", "pixel lon"),
        "pixel_lat": pick("lat", "latitude", "pixel lat"),
        "pixel_id": pick("Pixel_ID", "pixelid"),
        "index_id": pick("Index_ID", "indexid", "index"),
        "farmer_count": pick("Farmer Number", "farmercount", "farmers", "n_farmers"),
    }

def _first_non_null(s: pd.Series):
    s2 = s.dropna()
    return s2.iloc[0] if not s2.empty else None

def autosize_columns(ws, col_start: int, col_end: int, min_width: int = 8, max_width: int = 40):
    for col in range(col_start, col_end + 1):
        max_len = 0
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=1, max_row=ws.max_row, values_only=True):
            for v in cell:
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, min(max_width, max_len + 2))

def ws_title_is_default(title: str) -> bool:
    return str(title).lower().startswith("sheet")

def build_excel_sheet1(
    df: pd.DataFrame,
    wb: Optional[Workbook] = None,
    sheet_name: str = "1. Modelled Yield"
) -> Workbook:
    cols = _resolve_cols(df)
    required = {k: cols[k] for k in ("pixel_key", "year", "yield")}
    missing = [k for k, v in required.items() if not v]
    if missing:
        raise ValueError(f"Missing required columns: {missing}. Available: {list(df.columns)}")

    pixel_col = cols["pixel_key"]
    year_col = cols["year"]
    yield_col = cols["yield"]

    # Clean year and build full year list
    yr = pd.to_numeric(df[year_col], errors="coerce")
    df = df.assign(**{year_col: yr})
    all_years = (
        pd.Series(df[year_col].dropna().astype(int).unique())
        .sort_values()
        .tolist()
    )

    # Pivot (keep all years, including all-blank)
    pivot = (
        df.pivot_table(index=year_col, columns=pixel_col, values=yield_col, aggfunc="first")
        .sort_index()
        .reindex(all_years)
    )

    pixel_order: List = [c for c in list(pivot.columns) if not pd.isna(c)]

    # Per-pixel metadata
    meta: Dict[object, Dict[str, Optional[object]]] = {}
    for pix in pixel_order:
        sub = df[df[pixel_col] == pix]
        lon = _first_non_null(sub[cols["pixel_lon"]]) if cols["pixel_lon"] else None
        lat = _first_non_null(sub[cols["pixel_lat"]]) if cols["pixel_lat"] else None
        meta[pix] = {
            "area": _first_non_null(sub[cols["area"]]) if cols["area"] else None,
            "region": _first_non_null(sub[cols["region"]]) if cols["region"] else None,
            "indexid": _first_non_null(sub[cols["index_id"]]) if cols.get("index_id") else None,
            "farmer_count": _first_non_null(sub[cols["farmer_count"]]) if cols.get("farmer_count") else None,
            "lon": float(lon) if lon is not None and pd.notna(lon) else None,
            "lat": float(lat) if lat is not None and pd.notna(lat) else None,
            "pixelid": _first_non_null(sub[cols["pixel_id"]]) if cols["pixel_id"] else None,
        }

    # Workbook/sheet setup
    wb = wb or Workbook()
    ws = wb.active if (wb.active and wb.active.max_row == 1 and ws_title_is_default(wb.active.title)) else wb.create_sheet()
    ws.title = sheet_name
    ws.freeze_panes = ws.cell(row=10, column=6)  # F10

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    left = Alignment(horizontal="left")

    col_label = 5          # E
    first_data_col = 6     # F
    row_title = 2
    row_meta_start = 3
    row_data_start = 10

    # Title
    ws.cell(row=row_title, column=col_label).value = "MODELLED YIELDS (tons per ha)"
    ws.cell(row=row_title, column=col_label).font = bold
    ws.cell(row=row_title, column=col_label).alignment = left
    ws.cell(row=row_title, column=col_label).fill = to_fill("FFFF00")

    # Pixel count
    ws.cell(row=row_meta_start + 0, column=col_label).value = "Pixel count"
    for j, _pix in enumerate(pixel_order):
        c = ws.cell(row=row_meta_start + 0, column=first_data_col + j)
        c.value = j + 1
        c.alignment = center

    # Area (now with color fill)
    ws.cell(row=row_meta_start + 1, column=col_label).value = "Area"
    for j, pix in enumerate(pixel_order):
        cell = ws.cell(row=row_meta_start + 1, column=first_data_col + j)
        area_name = meta[pix]["area"] or ""
        cell.value = area_name
        # match color case-insensitively
        hex6 = None
        if area_name:
            for k, v in AREA_COLORS_HEX.items():
                if str(area_name).strip().lower() == k.lower():
                    hex6 = v
                    break
        if hex6:
            cell.fill = to_fill(hex6)

    # Region
    ws.cell(row=row_meta_start + 2, column=col_label).value = "Region"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=row_meta_start + 2, column=first_data_col + j).value = meta[pix]["region"] or ""

    # Farmer count (replaces old "Index ID" row)
    ws.cell(row=row_meta_start + 3, column=col_label).value = "Farmer count"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=row_meta_start + 3, column=first_data_col + j).value = meta[pix]["farmer_count"]

    # Pixel Lon
    ws.cell(row=row_meta_start + 4, column=col_label).value = "Pixel Lon"
    for j, pix in enumerate(pixel_order):
        v = meta[pix]["lon"]
        ws.cell(row=row_meta_start + 4, column=first_data_col + j).value = float(v) if isinstance(v, (int, float)) and not math.isnan(v) else v

    # Note + Pixel Lat
    red_font = Font(color="FF0000")
    note_cell = ws.cell(row=row_meta_start + 5, column=1)
    note_cell.value = "Note: if yield data absent, then pixel has dropped out"
    note_cell.font = red_font

    ws.cell(row=row_meta_start + 5, column=col_label).value = "Pixel Lat"
    for j, pix in enumerate(pixel_order):
        v = meta[pix]["lat"]
        ws.cell(row=row_meta_start + 5, column=first_data_col + j).value = float(v) if isinstance(v, (int, float)) and not math.isnan(v) else v

    # Pixel ID
    ws.cell(row=row_meta_start + 6, column=col_label).value = "Pixel ID"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=row_meta_start + 6, column=first_data_col + j).value = meta[pix]["pixelid"] or pix

    # Data rows
    years = list(pivot.index)
    for i, y in enumerate(years, start=1):
        r = row_data_start + i - 1
        # row counter in col D
        c_idx = ws.cell(row=r, column=col_label - 1)
        c_idx.value = i
        c_idx.alignment = center
        # Year in col E
        try:
            ws.cell(row=r, column=col_label).value = int(y)
        except Exception:
            ws.cell(row=r, column=col_label).value = y
        # Yields across pixels
        for j in range(len(pixel_order)):
            val = pivot.iat[i - 1, j]
            ws.cell(row=r, column=first_data_col + j).value = (float(val) if pd.notnull(val) else None)

    # Bold labels in column E
    for rr in range(row_meta_start, row_meta_start + 7):
        c = ws.cell(row=rr, column=col_label)
        c.font = bold
        c.alignment = left

    last_col = max(first_data_col + len(pixel_order) - 1, col_label)
    autosize_columns(ws, 1, last_col)



    # === BEGIN: Formatting tweaks per request (v2) ===

    from openpyxl.styles import Alignment as _Align1

    from openpyxl.utils import get_column_letter as _gcl1


    # Task 4: Merge title with neighbor on the right; set column E width -> 18.0

    try:

        ws.merge_cells(start_row=row_title, start_column=col_label, end_row=row_title, end_column=first_data_col)

    except Exception:

        pass

    ws.column_dimensions['E'].width = 18.0


    # Task 1: Set data columns (F and onward) width -> 21.4

    last_data_col = first_data_col + len(pixel_order) - 1

    for _c in range(first_data_col, max(first_data_col, last_data_col) + 1):

        ws.column_dimensions[_gcl1(_c)].width = 21.4


    # Task 6: Columns A-D width -> 7.0; merge A8:D9 and apply wrap text

    for _col in ['A','B','C','D']:

        ws.column_dimensions[_col].width = 7.0

    try:

        ws.merge_cells(start_row=8, start_column=1, end_row=9, end_column=4)

        _cell = ws.cell(row=8, column=1)

        _cell.alignment = _Align1(wrap_text=True, horizontal=_cell.alignment.horizontal if _cell.alignment else None, vertical=_cell.alignment.vertical if _cell.alignment else None)

    except Exception:

        pass

    # === END: Formatting tweaks per request (v2) ===


    return wb
