import pandas as pd
from typing import Optional, Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

def _norm(s: str) -> str:
    return str(s).strip().lower().replace(" ", "").replace("_", "")

def _resolve_cols(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    norm_to_orig = {_norm(c): c for c in df.columns}
    def pick(*aliases: str) -> Optional[str]:
        for a in aliases:
            k = _norm(a)
            if k in norm_to_orig:
                return norm_to_orig[k]
        return None
    return {
        "pixel_key": pick("Pixel_ID", "pixelid", "pixel"),
        "year":      pick("Year", "year"),
        "loan":      pick("Loan_Amount", "loanamount", "loan"),
        "area":      pick("Area", "area_ha", "hectares"),
        "region":    pick("Region"),
        "lon":       pick("lon", "longitude"),
        "lat":       pick("lat", "latitude"),
        "pixel_id":  pick("Pixel_ID", "pixelid"),
    }

def _first_non_null(s: pd.Series):
    s2 = s.dropna()
    return s2.iloc[0] if not s2.empty else None

def _fill(hex6: str) -> PatternFill:
    return PatternFill(fill_type="solid", start_color=f"FF{hex6}", end_color=f"FF{hex6}")

def _autosize(ws, c1: int, c2: int, min_w: int = 8, max_w: int = 40):
    for col in range(c1, c2 + 1):
        m = 0
        for vals in ws.iter_cols(min_col=col, max_col=col, min_row=1, max_row=ws.max_row, values_only=True):
            for v in vals:
                if v is None:
                    continue
                m = max(m, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = max(min_w, min(max_w, m + 2))

def ws_title_is_default(title: str) -> bool:
    return str(title).lower().startswith("sheet")

def build_excel_sheet3(
    df: pd.DataFrame,
    wb: Optional[Workbook] = None,
    sheet_name: str = "3. Payout Amounts"
) -> Workbook:
    """
    Sheet 3: Payout Amounts (USD)
    - Starts at row 1 (title), metadata rows 2..8, headers row 9, data row 10.
    - Summary tweaks per request:
        A3: "Total Loan Amounts (USD)"
        C3: value = SUM(loan row)   [change to C5 if you truly want both values in C5]
        A5: "Total Number of Pixels"
        C5: value = number of pixel columns
        D1:D3 (merged, red): "Note: Loan amounts are given as average loan amount across pixels within a given region,
                              as pixel-level loan amount info not available yet."
    - Metadata:
        Row 2: Pixel count (values 1..N)
        Row 3: Loan Amounts (USD) [values]
        Row 4: Sum Insured (USD)  [FORMULA = 0.4 * Loan]
        Row 5: Area [values]
        Row 6: Region [values]
        Row 7: Pixel Lon [values]
        Row 8: Pixel Lat [values]
    - Grid: payout_amount = ('2. Payouts %'!rc) * (this sheet’s Sum Insured for the column), blank-safe.
    """
    cols = _resolve_cols(df)
    if not cols["pixel_key"] or not cols["year"]:
        raise ValueError("Dataframe must include Pixel_ID and Year.")

    pixel_col = cols["pixel_key"]; year_col = cols["year"]
    pixel_order = sorted(df[pixel_col].dropna().unique().tolist())
    year_list   = sorted(df[year_col].dropna().unique().tolist())

    # Per-pixel metadata (values; Sum Insured will be a formula)
    meta: Dict[object, Dict[str, Optional[object]]] = {}
    for pix in pixel_order:
        sub = df[df[pixel_col] == pix]
        meta[pix] = {
            "loan":    _first_non_null(sub[cols["loan"]])   if cols["loan"]   else None,
            "area":    _first_non_null(sub[cols["area"]])   if cols["area"]   else None,
            "region":  _first_non_null(sub[cols["region"]]) if cols["region"] else None,
            "lon":     _first_non_null(sub[cols["lon"]])    if cols["lon"]    else None,
            "lat":     _first_non_null(sub[cols["lat"]])    if cols["lat"]    else None,
            "pixelid": _first_non_null(sub[cols["pixel_id"]]) if cols["pixel_id"] else pix,
        }

    wb = wb or Workbook()
    ws = wb.active if (wb.active and wb.active.max_row == 1 and ws_title_is_default(wb.active.title)) else wb.create_sheet()
    ws.title = sheet_name

    bold = Font(bold=True)
    red_font = Font(color="FF0000")
    left = Alignment(horizontal="left")
    center = Alignment(horizontal="center")

    # Layout constants
    COL_YEAR_LABEL  = 5   # E
    COL_FIRST_PIXEL = 6   # F
    ROW_TITLE       = 1
    ROW_META_START  = 2   # rows 2..8 metadata
    ROW_PIXEL_ID    = 9   # header row
    ROW_FIRST_DATA  = 10  # first data row
    ws.freeze_panes = ws.cell(row=ROW_FIRST_DATA, column=COL_FIRST_PIXEL)

    # Title
    ws.cell(row=ROW_TITLE, column=COL_YEAR_LABEL).value = "PAYOUT AMOUNTS (USD)"
    ws.cell(row=ROW_TITLE, column=COL_YEAR_LABEL).font = bold
    ws.cell(row=ROW_TITLE, column=COL_YEAR_LABEL).alignment = left
    ws.cell(row=ROW_TITLE, column=COL_YEAR_LABEL).fill = _fill("FFFF00")

    # ===== Summary (exact placement per request) =====
    # Labels
    ws.cell(row=3, column=1).value = "Total Loan Amounts (USD)"
    ws.cell(row=3, column=1).font = bold
    ws.cell(row=5, column=1).value = "Total Number of Pixels"
    ws.cell(row=5, column=1).font = bold

    # Compute ranges for summary values
    last_col_idx = COL_FIRST_PIXEL + len(pixel_order) - 1
    first_col_letter = get_column_letter(COL_FIRST_PIXEL)
    last_col_letter  = get_column_letter(last_col_idx)
    loans_row = ROW_META_START + 1  # row 3 (loan)
    loans_rng = f"{first_col_letter}{loans_row}:{last_col_letter}{loans_row}"

    # Values (using C3 and C5)
    row_val_total_loans = 3
    ws.cell(row=row_val_total_loans, column=3).value = f"=IF(COUNT({loans_rng})=0,\"\",SUM({loans_rng}))"
    ws.cell(row=row_val_total_loans, column=3).number_format = "#,##0"
    first_col_letter = get_column_letter(COL_FIRST_PIXEL)
    last_col_letter  = get_column_letter(COL_FIRST_PIXEL + len(pixel_order) - 1)
    loan_row_range = f"${first_col_letter}$3:${last_col_letter}$3"
    ws.cell(row=5, column=3).value = f"=COUNT({loan_row_range})"

    # Note in D1:D3 (merged, red)
    ws.merge_cells(start_row=1, start_column=4, end_row=3, end_column=4)
    note_cell = ws.cell(row=1, column=4)
    note_cell.value = ("Note: Loan amounts are given as average loan amount across pixels within a given region, "
                       "as pixel-level loan amount info not available yet.")
    note_cell.font = red_font
    note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ===== Metadata labels (E), values F→ =====
    ws.cell(row=ROW_META_START + 0, column=COL_YEAR_LABEL).value = "Pixel count"
    ws.cell(row=ROW_META_START + 1, column=COL_YEAR_LABEL).value = "Loan Amounts (USD)"
    ws.cell(row=ROW_META_START + 2, column=COL_YEAR_LABEL).value = "Sum Insured (USD)"
    ws.cell(row=ROW_META_START + 3, column=COL_YEAR_LABEL).value = "Area"
    ws.cell(row=ROW_META_START + 4, column=COL_YEAR_LABEL).value = "Region"
    ws.cell(row=ROW_META_START + 5, column=COL_YEAR_LABEL).value = "Pixel Lon"
    ws.cell(row=ROW_META_START + 6, column=COL_YEAR_LABEL).value = "Pixel Lat"

    area_colors_hex = {
        "Northern Zone": "1F77B4", "Central Zone": "2CA02C", "Lake Zone": "FF7F0E",
        "Western Zone": "9467BD", "Southern Highlands Zone": "8C564B",
        "Coastal Zone": "17BECF", "Zanzibar (Islands)": "7F7F7F",
    }

    for j, pix in enumerate(pixel_order):
        c = COL_FIRST_PIXEL + j
        ws.cell(row=ROW_META_START + 0, column=c).value = j + 1  # Pixel count ordinal

        # Loan (value)
        ws.cell(row=ROW_META_START + 1, column=c).value = meta[pix]["loan"]

        # Sum Insured (formula = 0.4*Loan, blank-safe)
        loan_ref = f"{get_column_letter(c)}{ROW_META_START + 1}"
        ws.cell(row=ROW_META_START + 2, column=c).value = f"=IF({loan_ref}=\"\",\"\",0.4*{loan_ref})"
        ws.cell(row=ROW_META_START + 2, column=c).number_format = "#,##0"

        # Area (colored if matched)
        area_cell = ws.cell(row=ROW_META_START + 3, column=c)
        v_area = meta[pix]["area"] or ""
        area_cell.value = v_area
        hx = area_colors_hex.get(str(v_area))
        if not hx and v_area:
            for k, val in area_colors_hex.items():
                if str(v_area).strip().lower() == k.lower():
                    hx = val; break
        if hx:
            area_cell.fill = _fill(hx)

        ws.cell(row=ROW_META_START + 4, column=c).value = meta[pix]["region"] or ""
        ws.cell(row=ROW_META_START + 5, column=c).value = meta[pix]["lon"]
        ws.cell(row=ROW_META_START + 6, column=c).value = meta[pix]["lat"]

    # ===== Header row (row 9) =====
    ws.cell(row=ROW_PIXEL_ID, column=2).value = "SD"
    ws.cell(row=ROW_PIXEL_ID, column=4).value = "Average"
    ws.cell(row=ROW_PIXEL_ID, column=COL_YEAR_LABEL).value = "Year"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=ROW_PIXEL_ID, column=COL_FIRST_PIXEL + j).value = meta[pix]["pixelid"] or str(pix)

    # Year labels (E10..)
    for i, y in enumerate(year_list):
        r = ROW_FIRST_DATA + i
        try:
            ws.cell(row=r, column=COL_YEAR_LABEL).value = int(y)
        except Exception:
            ws.cell(row=r, column=COL_YEAR_LABEL).value = y

    # ===== GRID: payout amount = Sheet2% * SumInsured (blank-safe) =====
    sheet2_name = "2. Payouts %"
    last_col_idx = COL_FIRST_PIXEL + len(pixel_order) - 1
    last_col_letter = get_column_letter(last_col_idx)

    for i, _ in enumerate(year_list):
        r = ROW_FIRST_DATA + i
        row_rng = f"{get_column_letter(COL_FIRST_PIXEL)}{r}:{last_col_letter}{r}"

        # Per-year stats (A–D), blank-safe
        ws.cell(row=r, column=2).value = f"=IF(COUNT({row_rng})<=1,\"\",STDEV({row_rng}))"
        ws.cell(row=r, column=4).value = f"=IF(COUNT({row_rng})=0,\"\",AVERAGE({row_rng}))"
        ws.cell(row=r, column=2).number_format = "#,##0"
        ws.cell(row=r, column=4).number_format = "#,##0"

        # Grid cells
        for j, _ in enumerate(pixel_order):
            c = COL_FIRST_PIXEL + j
            colL = get_column_letter(c)
            pct_ref = f"'{sheet2_name}'!{colL}{r}"
            si_ref  = f"{colL}{ROW_META_START + 2}"  # Sum Insured row
            formula = (
                f"=IF(OR(ISBLANK({pct_ref}),NOT(ISNUMBER({pct_ref}))),\"\","
                f"IF(OR(ISBLANK({si_ref}),NOT(ISNUMBER({si_ref}))),\"\","
                f"{pct_ref}*{si_ref}))"
            )
            cell = ws.cell(row=r, column=c)
            cell.value = formula
            cell.number_format = "#,##0"

    # ===== Styling & sizing =====
    for rr in range(ROW_META_START, ROW_PIXEL_ID + 1):
        ws.cell(row=rr, column=COL_YEAR_LABEL).font = bold
        ws.cell(row=rr, column=COL_YEAR_LABEL).alignment = left
    for cc in [2, 4]:
        ws.cell(row=ROW_PIXEL_ID, column=cc).font = bold
        ws.cell(row=ROW_PIXEL_ID, column=cc).alignment = center
    for j in range(len(pixel_order)):
        ws.cell(row=ROW_PIXEL_ID, column=COL_FIRST_PIXEL + j).font = bold
        ws.cell(row=ROW_PIXEL_ID, column=COL_FIRST_PIXEL + j).alignment = center

    last_col = max(COL_FIRST_PIXEL + len(pixel_order) - 1, COL_YEAR_LABEL)
    _autosize(ws, 1, last_col)

    # Recalc on open (safe)
    try:
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    except Exception:
        try:
            wb.calculation.fullCalcOnLoad = True
        except Exception:
            pass

    return wb
