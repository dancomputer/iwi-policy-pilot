from typing import Optional, Sequence, Union
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pyparsing import Path

# Import your existing builder
from MakeExogenousExcelInputDataframe import build_regional_statistics

DASH = "-   "   # placeholder (dash + spaces like your sample)

ROW_ORDER_DEFAULT: Sequence[str] = (
    "Loan amounts (USD)",
    "Sum insured",
    "Area",
    "Region",
    # years inserted dynamically after these two
    "Average Payout",
    "SD",
    "Min",
    "Max",
    "90th percentile",
    "95th percentile",
    "Number of Pixels",
    "Number of Zero and Blank Pixels",
    "Number of Blank Pixels",
    "Number of Zero Pixel",
    "Average non-zero/blank pixel CoV",
)

INT_ROWS = {
    "Loan amounts (USD)",
    "Sum insured",
    "Average Payout",
    "SD",
    "Min",
    "Max",
    "90th percentile",
    "95th percentile",
    "Number of Pixels",
    "Number of Zero and Blank Pixels",
    "Number of Blank Pixels",
    "Number of Zero Pixel",
}

COV_ROW = "Average non-zero/blank pixel CoV"

# Area color mapping (same as sheet 4)
AREA_COLORS_HEX = {
    "Northern Zone": "1F77B4",
    "Central Zone": "2CA02C",
    "Lake Zone": "FF7F0E",
    "Western Zone": "9467BD",
    "Southern Highlands Zone": "8C564B",
    "Coastal Zone": "17BECF",
    "Zanzibar (Islands)": "7F7F7F",
}

def _space_thousands(n):
    try:
        return f"{int(round(float(n))):,}".replace(",", " ")
    except Exception:
        return DASH

def _format_cov(x: Optional[float]) -> str:
    if x is None or pd.isna(x):
        return DASH
    return f"{float(x):.2f}".replace(".", ",")

def _detect_year_rows(df_numeric: pd.DataFrame) -> Sequence[str]:
    years = []
    for idx in df_numeric.index:
        if idx.isdigit() and 1900 <= int(idx) <= 2100:
            years.append(idx)
    return sorted(years)

def _auto_size(ws: Worksheet, min_width=6, max_width=32):
    for col in range(1, ws.max_column + 1):
        mx = 0
        for cell in ws.iter_cols(min_col=col, max_col=col,
                                 min_row=1, max_row=ws.max_row, values_only=True):
            for v in cell:
                if v is None:
                    continue
                ln = len(str(v))
                if ln > mx:
                    mx = ln
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, min(max_width, mx + 2))

def _area_fill(area_name: str) -> Optional[PatternFill]:
    if not area_name:
        return None
    for k, hexv in AREA_COLORS_HEX.items():
        if str(area_name).strip().lower() == k.lower():
            return PatternFill(fill_type="solid", start_color=f"FF{hexv}", end_color=f"FF{hexv}")
    return None

def build_excel_sheet5(
    df_wide_numeric: pd.DataFrame,
    df_wide_formatted: pd.DataFrame,
    wb: Optional[Workbook] = None,
    sheet_name: str = "5. Regional Statistics",
    merge_overall_total: bool = True,
) -> Workbook:
    # Validate shape
    if not df_wide_numeric.index.equals(df_wide_formatted.index):
        raise ValueError("Index mismatch between numeric and formatted wide dataframes.")
    if list(df_wide_numeric.columns) != list(df_wide_formatted.columns):
        raise ValueError("Column mismatch between numeric and formatted wide dataframes.")

    wb = wb or Workbook()
    if (wb.active and wb.active.max_row == 1 and wb.active.max_column == 1
            and wb.active.title.lower().startswith("sheet")):
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)

    bold = Font(bold=True)
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    # Determine ordered rows including years just before first statistics row
    year_rows = _detect_year_rows(df_wide_numeric)
    ordered_rows = []
    for r in ROW_ORDER_DEFAULT:
        if r == "Average Payout":
            ordered_rows.extend(year_rows)
        ordered_rows.append(r)
    remaining = [r for r in df_wide_numeric.index if r not in ordered_rows]
    ordered_rows.extend(remaining)

    # Rows whose label (first column) should NOT be bold
    NOT_BOLD_ROWS = {
        "Number of Zero and Blank Pixels",
        "Number of Blank Pixels",
        "Number of Zero Pixel",
        "Average non-zero/blank pixel CoV",
    }
    # Add the dynamic year rows to NOT_BOLD_ROWS
    NOT_BOLD_ROWS.update(year_rows)

    # Write table
    for row_idx_excel, row_label in enumerate(ordered_rows, start=1):
        label_cell = ws.cell(row=row_idx_excel, column=1, value=row_label)
        # Only bold if not in NOT_BOLD_ROWS
        if row_label not in NOT_BOLD_ROWS:
            label_cell.font = bold
        label_cell.alignment = left

        for col_offset, col_label in enumerate(df_wide_numeric.columns, start=2):
            raw_val = df_wide_numeric.at[row_label, col_label] if row_label in df_wide_numeric.index else None
            excel_cell = ws.cell(row=row_idx_excel, column=col_offset)

            if row_label == "Area":
                val = df_wide_formatted.at[row_label, col_label]
                if val is None or pd.isna(val) or str(val).strip() == "":
                    excel_cell.value = DASH
                else:
                    excel_cell.value = val
                    # Conditional coloring by area
                    fill = _area_fill(str(val))
                    if fill:
                        excel_cell.fill = fill
                excel_cell.alignment = center
                continue

            if row_label == "Region":
                val = df_wide_formatted.at[row_label, col_label]
                if val is None or pd.isna(val) or str(val).strip() == "":
                    excel_cell.value = DASH
                else:
                    excel_cell.value = val
                excel_cell.alignment = center
                continue

            if row_label == COV_ROW:
                if raw_val is None or pd.isna(raw_val):
                    excel_cell.value = DASH
                else:
                    excel_cell.value = _format_cov(raw_val)
                excel_cell.alignment = right
                continue

            if row_label in INT_ROWS or row_label in year_rows:
                if raw_val is None or pd.isna(raw_val) or (
                    isinstance(raw_val, (int, float)) and float(raw_val) == 0
                ):
                    excel_cell.value = DASH
                else:
                    try:
                        ival = int(round(float(raw_val)))
                        excel_cell.value = ival
                        excel_cell.number_format = "# ##0"
                    except Exception:
                        excel_cell.value = DASH
                excel_cell.alignment = right
                continue

            if raw_val is None or pd.isna(raw_val):
                excel_cell.value = DASH
            else:
                excel_cell.value = str(raw_val)
            excel_cell.alignment = right

    # Merge "Overall Total" vertically across Area + Region rows if requested
    if merge_overall_total and "Area" in ordered_rows and "Region" in ordered_rows:
        area_row_num = ordered_rows.index("Area") + 1
        region_row_num = ordered_rows.index("Region") + 1
        for col_offset, col_label in enumerate(df_wide_numeric.columns, start=2):
            area_cell_val = ws.cell(row=area_row_num, column=col_offset).value
            if area_cell_val and str(area_cell_val).strip().lower() == "overall total":
                ws.cell(row=region_row_num, column=col_offset).value = None
                ws.merge_cells(start_row=area_row_num, start_column=col_offset,
                               end_row=region_row_num, end_column=col_offset)
                merged = ws.cell(row=area_row_num, column=col_offset)
                merged.alignment = center
                merged.font = bold
                break

    # Freeze panes below Region row, after first column
    if "Region" in ordered_rows:
        freeze_row = ordered_rows.index("Region") + 2
        ws.freeze_panes = ws.cell(row=freeze_row, column=2)

    ws.column_dimensions[get_column_letter(1)].width = 36
    _auto_size(ws, max_width=22)

    return wb

def build_sheet5_from_df_final(
    df_final: pd.DataFrame,
    wb: Optional[Workbook] = None,
    sheet_name: str = "5. Regional Totals",
    verbose: bool = False
) -> Workbook:
    df_wide_numeric, df_wide_formatted = build_regional_statistics(df_final, verbose=verbose)
    return build_excel_sheet5(df_wide_numeric, df_wide_formatted, wb=wb, sheet_name=sheet_name)

def save_sheet5_from_regional_dataframes(
    df_wide_numeric: pd.DataFrame,
    df_wide_formatted: pd.DataFrame,
    file_path: Union[str, Path],
    sheet_name: str = "5. Regional Totals",
) -> str:
    """
    Build '5. Regional Totals' from the provided regional dataframes and save to file_path.

    Parameters
    - df_wide_numeric: numeric wide dataframe returned by build_regional_statistics
    - df_wide_formatted: formatted wide dataframe returned by build_regional_statistics
    - file_path: destination .xlsx path
    - sheet_name: optional sheet name override

    Returns
    - The saved file path as a string
    """
    p = Path(file_path)
    p.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb = build_excel_sheet5(df_wide_numeric, df_wide_formatted, wb=wb, sheet_name=sheet_name)
    wb.save(p)
    return str(p)

if __name__ == "__main__":
    print("Module ready: build_excel_sheet5 with area coloring and adjusted bold rules.")