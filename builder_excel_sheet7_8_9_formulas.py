# builder_excel_sheet7_8.py

from typing import Optional, List, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
# Keep area palette consistent with your other sheets
AREA_ORDER = [
    "Northern Zone",
    "Central Zone",
    "Lake Zone",
    "Western Zone",
    "Southern Highlands Zone",
    "Coastal Zone",
    "Zanzibar (Islands)",
]
AREA_COLORS_HEX = {
    "Northern Zone": "1F77B4",
    "Central Zone": "2CA02C",
    "Lake Zone": "FF7F0E",
    "Western Zone": "9467BD",
    "Southern Highlands Zone": "8C564B",
    "Coastal Zone": "17BECF",
    "Zanzibar (Islands)": "7F7F7F",
}

def _detect_years(df: pd.DataFrame) -> List[int]:
    yrs = []
    for idx in df.index:
        s = str(idx)
        if s.isdigit():
            y = int(s)
            if 1900 <= y <= 2100:
                yrs.append(y)
    return sorted(set(yrs))

def _areas_present_on_sheet3(wb: Workbook, sheet3_name: str = "3. Payout Amounts") -> List[str]:
    """
    Inspect Sheet 3 to find which areas actually exist:
      - Pixel headers are on row 9 starting at column F
      - Area names are on row 5 aligned to those pixel columns
    Returns unique area list in the order of AREA_ORDER.
    """
    if sheet3_name not in [ws.title for ws in wb.worksheets]:
        # If Sheet 3 isn't there yet, fall back to the full list (safe default)
        return list(AREA_ORDER)

    ws3 = wb[sheet3_name]
    row_header = 9  # pixel header row
    row_area   = 5  # area names row

    # Find pixel columns: non-empty cells in row 9 from col F onward
    start_col_idx = 6  # column F = 6 (1-based)
    pixel_cols = []
    for col in range(start_col_idx, ws3.max_column + 1):
        if ws3.cell(row=row_header, column=col).value not in (None, ""):
            pixel_cols.append(col)

    # Collect unique areas from row 5 at those pixel columns
    found = set()
    ordered = []
    for col in pixel_cols:
        val = ws3.cell(row=row_area, column=col).value
        name = (str(val).strip() if val is not None else "")
        if name:
            found.add(name)

    # Keep only areas that exist, preserving your preferred AREA_ORDER
    for a in AREA_ORDER:
        if a in found:
            ordered.append(a)

    # As a fallback, if nothing detected but headers exist, include whatever non-empty unique names we saw
    if not ordered and found:
        ordered = sorted(found)

    # If still nothing (truly empty), return empty list (Sheet 7 will still write header row only)
    return ordered

def build_excel_sheet7_chartdata(
    df_wide_numeric: pd.DataFrame,
    wb: Optional[Workbook] = None,
    sheet_name: str = "7. Chart Data",
) -> Workbook:
    """
    Build Sheet 7 with a formula-driven table for the Area Payout chart.
    Columns: Year, [Area 1], [Area 2], ...
    Each data cell is blank-safe: =IF(count=0,"",sum) where
      count = SUMPRODUCT(--(AreaRow="Area"), --ISNUMBER(YearRow))
      sum   = SUMPRODUCT(--(AreaRow="Area"), YearRow)
    """
    wb = wb or Workbook()

    # Create / reuse the sheet
    if sheet_name in [ws.title for ws in wb.worksheets]:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=sheet_name[:31])

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    right  = Alignment(horizontal="right")

    years = _detect_years(df_wide_numeric)
    first_year = years[0] if years else None

    # Only include areas actually present on Sheet 3
    areas_present = _areas_present_on_sheet3(wb, sheet3_name="3. Payout Amounts")

    # Headers
    ws.cell(row=1, column=1, value="Year").font = bold
    for j, area in enumerate(areas_present, start=2):
        ws.cell(row=1, column=j, value=area).font = bold

    # Dynamic width across pixel columns on Sheet 3 (row 9 has pixel headers starting at F)
    HCOUNT = "COUNTA('3. Payout Amounts'!F9:XFD9)"
    def row_range_on_sheet3(row_num: int) -> str:
        # 1×N horizontal range starting at F<row>, width = HCOUNT
        return f"OFFSET('3. Payout Amounts'!F{row_num},0,0,1,{HCOUNT})"

    AREA_ROW_RANGE = row_range_on_sheet3(5)  # area names in row 5

    # Rows (blank-safe: show "" when an area has no data that year)
    if years and areas_present:
        for i, y in enumerate(years, start=2):
            ws.cell(row=i, column=1, value=y).alignment = center

            sheet3_row_for_y = 10 + (y - first_year)  # Sheet 3 payouts start at row 10
            SUM_ROW_Y = row_range_on_sheet3(sheet3_row_for_y)

            for j, area in enumerate(areas_present, start=2):
                count_expr = f"SUMPRODUCT(--({AREA_ROW_RANGE}=\"{area}\"),--ISNUMBER({SUM_ROW_Y}))"
                sum_expr   = f"SUMPRODUCT(--({AREA_ROW_RANGE}=\"{area}\"),{SUM_ROW_Y})"
                formula    = f"=IF({count_expr}=0,\"\",{sum_expr})"
                c = ws.cell(row=i, column=j)
                c.value = formula
                c.alignment = right
                c.number_format = "# ##0"

    # Freeze + widths
    ws.freeze_panes = ws["B2"]
    ws.column_dimensions["A"].width = 10
    for col in range(2, 2 + len(areas_present)):
        ws.column_dimensions[get_column_letter(col)].width = 18

    return wb

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties,RichTextProperties

# Consistent area palette (ARGB = "FF" + HEX)
AREA_COLORS_HEX = {
    "Northern Zone": "1F77B4",
    "Central Zone": "2CA02C",
    "Lake Zone": "FF7F0E",
    "Western Zone": "9467BD",
    "Southern Highlands Zone": "8C564B",
    "Coastal Zone": "17BECF",
    "Zanzibar (Islands)": "7F7F7F",
}
from openpyxl.chart.layout import Layout, ManualLayout

def build_excel_sheet8_area_payout_chart(
    wb,
    sheet7_name: str = "7. Chart Data",
    sheet8_name: str = "8. Area Payout Chart",
):
    if sheet7_name not in [ws.title for ws in wb.worksheets]:
        raise ValueError(f"{sheet7_name} must exist before creating the chart.")
    ws7 = wb[sheet7_name]

    # Last used Year row
    last_row = 1
    for r in range(ws7.max_row, 1, -1):
        if ws7.cell(row=r, column=1).value not in (None, ""):
            last_row = r
            break
    if last_row < 2:
        last_row = 2

    # Detect ORIGINAL area block: B..first blank header
    first_area_col = 2
    last_area_col = first_area_col - 1
    j = first_area_col
    while True:
        header = ws7.cell(row=1, column=j).value
        if header in (None, ""):
            break
        last_area_col = j
        j += 1
    if last_area_col < first_area_col:
        # nothing to chart
        if sheet8_name in [ws.title for ws in wb.worksheets]:
            wb.remove(wb[sheet8_name])
        ws8 = wb.create_sheet(title=sheet8_name[:31])
        ws8["B3"] = "No area payout data available to chart."
        return wb

    # Recreate Sheet 8 fresh
    if sheet8_name in [ws.title for ws in wb.worksheets]:
        wb.remove(wb[sheet8_name])
    ws8 = wb.create_sheet(title=sheet8_name[:31])

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.gapWidth = 40
    chart.title = "Area Payout (US Dollars) per Year"

    # 1) Title should NOT overlay the plot
    chart.title.overlay = False

    # 2) Nudge the plot area down a bit so the (bigger) title has room
    #    (x/y/w/h are fractions of the chart frame; tweak to taste)
    #chart.layout = Layout(
    #    manualLayout=ManualLayout(
    #        x=0.06,   # left margin ~6%
    #        y=0.14,   # top margin ~14% (increase if title still touches bars)
    #        w=0.88,   # width (1 - x - right margin)
    #        h=0.78,   # height (leave room for legend below)
    #    )
    #)
    
    chart.title.tx.rich.p[0].pPr.defRPr.sz = 2000
    chart.title.tx.rich.p[0].pPr.defRPr.b  = True

    chart.legend.position = "b"
    chart.legend.overlay = False
    chart.x_axis.title = None
    chart.y_axis.title = None
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.y_axis.number_format = "$#,##0"
    chart.y_axis.scaling.min = 0

    # Bigger tick labels + rotation (vertical here)
    AXIS_SZ = 1400
    LEGEND_SZ = 1300
    rotation_angle = -5400000  # -90° * 60000

    chart.x_axis.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=AXIS_SZ)),
                     endParaRPr=CharacterProperties())],
        bodyPr=RichTextProperties(rot=rotation_angle)
    )
    chart.y_axis.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=AXIS_SZ)),
                     endParaRPr=CharacterProperties())],
        bodyPr=RichTextProperties()
    )
    chart.legend.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=LEGEND_SZ)),
                     endParaRPr=CharacterProperties())],
        bodyPr=RichTextProperties()
    )

    # Add ONLY the original area series B..last_area_col; color safely with 6-digit RGB
    for j in range(first_area_col, last_area_col + 1):
        # Skip columns that are entirely blank in the data rows
        has_val = any(ws7.cell(row=r, column=j).value not in (None, "") for r in range(2, last_row + 1))
        if not has_val:
            continue
        series_ref = Reference(ws7, min_col=j, min_row=1, max_col=j, max_row=last_row)
        chart.add_data(series_ref, titles_from_data=True)

        header = str(ws7.cell(row=1, column=j).value or "")
        hex6 = AREA_COLORS_HEX.get(header)
        if hex6:
            try:
                s = chart.series[-1]
                s.graphicalProperties.solidFill = hex6  # safest
            except Exception:
                pass
    
    # Categories FIRST (A2:A<last_row>)
    cats = Reference(ws7, min_col=1, min_row=2, max_col=1, max_row=last_row)
    chart.set_categories(cats)

    # Keep bars centered over ticks; add slight side padding via layout if you want
    chart.y_axis.crossBetween = "between"

    chart.height = 18
    chart.width  = 32
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    ws8.add_chart(chart, "B3")
    return wb

def augment_sheet7_with_percentage_block(
    wb,
    sheet7_name: str = "7. Chart Data",
    sep: str = ","  # argument separator: ";" for EU locales, "," for US
):
    """
    Add a % block to the right of the existing totals on Sheet 7.

    Layout:
      A: Year
      B..?: original area totals (contiguous headers)
      (spacer): ONE blank column (so Sheet 8 won't pick this up)
      % block: SAME headers, each cell =
               IF(NOT(ISNUMBER(cell));"";IF(SUMPRODUCT(row_range)=0;"";cell/SUMPRODUCT(row_range)))

    All % cells are formulas and formatted as percent.
    """
    if sheet7_name not in [ws.title for ws in wb.worksheets]:
        raise ValueError(f"{sheet7_name} must exist before creating the percentage block.")
    ws7 = wb[sheet7_name]

    # last used Year row
    last_row = 1
    for r in range(ws7.max_row, 1, -1):
        if ws7.cell(row=r, column=1).value not in (None, ""):
            last_row = r
            break
    if last_row < 2:
        return wb

    # detect original area header block: B..first blank header
    first_area_col = 2
    last_area_col = first_area_col - 1
    j = first_area_col
    while True:
        header = ws7.cell(row=1, column=j).value
        if header in (None, ""):
            break
        last_area_col = j
        j += 1
    if last_area_col < first_area_col:
        return wb

    # spacer column (blank header)
    spacer_col = last_area_col + 1
    ws7.cell(row=1, column=spacer_col, value=None)

    # % block headers (same as totals)
    pct_start_col = spacer_col + 1
    for j_src in range(first_area_col, last_area_col + 1):
        ws7.cell(row=1, column=pct_start_col + (j_src - first_area_col),
                 value=ws7.cell(row=1, column=j_src).value)

    # build formulas
    first_letter = get_column_letter(first_area_col)
    last_letter  = get_column_letter(last_area_col)

    for r in range(2, last_row + 1):
        row_sum = f"SUMPRODUCT(${first_letter}{r}:${last_letter}{r})"
        for j_src in range(first_area_col, last_area_col + 1):
            src = f"{get_column_letter(j_src)}{r}"
            dst_col = pct_start_col + (j_src - first_area_col)

            # IF(NOT(ISNUMBER(src));"";IF(SUMPRODUCT($B$r:$<LAST>$r)=0;"";src/SUMPRODUCT($B$r:$<LAST>$r)))
            formula = (
                f"=IF(NOT(ISNUMBER({src})){sep}\"\"{sep}"
                f"IF({row_sum}=0{sep}\"\"{sep}{src}/{row_sum}))"
            )
            c = ws7.cell(row=r, column=dst_col)
            c.value = formula
            c.number_format = "0%"

    return wb
def build_excel_sheet9_area_payout_pct_chart(
    wb,
    sheet7_name: str = "7. Chart Data",
    sheet9_name: str = "9. Area Payout %",
    x_tick_rotation_deg: int = -45,
):
    """
    Build Sheet 9 as a 100% stacked chart from the percentage block on Sheet 7.
    It locates the first BLANK header after the original area block and takes the
    consecutive headers to the right as the % block.
    """
    if sheet7_name not in [ws.title for ws in wb.worksheets]:
        raise ValueError(f"{sheet7_name} must exist before creating the chart.")
    ws7 = wb[sheet7_name]

    # Last used Year row
    last_row = 1
    for r in range(ws7.max_row, 1, -1):
        if ws7.cell(row=r, column=1).value not in (None, ""):
            last_row = r
            break
    if last_row < 2:
        last_row = 2

    # Find original area block B..first blank header
    first_area_col = 2
    last_area_col = first_area_col - 1
    j = first_area_col
    while True:
        header = ws7.cell(row=1, column=j).value
        if header in (None, ""):
            break
        last_area_col = j
        j += 1
    if last_area_col < first_area_col:
        # nothing to chart
        if sheet9_name in [ws.title for ws in wb.worksheets]:
            wb.remove(wb[sheet9_name])
        ws9 = wb.create_sheet(title=sheet9_name[:31])
        ws9["B3"] = "No percentage data available to chart."
        return wb

    spacer_col = last_area_col + 1
    pct_start  = spacer_col + 1

    # Collect consecutive % headers to the right of pct_start
    pct_cols = []
    j = pct_start
    while True:
        header = ws7.cell(row=1, column=j).value
        if header in (None, ""):
            break
        # keep only if there's at least one non-blank value in rows 2..last_row
        has_val = any(ws7.cell(row=r, column=j).value not in (None, "") for r in range(2, last_row + 1))
        if has_val:
            pct_cols.append(j)
        j += 1

    if not pct_cols:
        if sheet9_name in [ws.title for ws in wb.worksheets]:
            wb.remove(wb[sheet9_name])
        ws9 = wb.create_sheet(title=sheet9_name[:31])
        ws9["B3"] = "No percentage data available to chart."
        return wb

    # Recreate Sheet 9
    if sheet9_name in [ws.title for ws in wb.worksheets]:
        wb.remove(wb[sheet9_name])
    ws9 = wb.create_sheet(title=sheet9_name[:31])

    # Build 100% stacked chart
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "percentStacked"
    chart.overlap = 100
    chart.gapWidth = 40
    chart.title = "Area Payout (% Share) per Year"
    # 1) Title should NOT overlay the plot
    chart.title.overlay = False

    chart.title.tx.rich.p[0].pPr.defRPr.sz = 2000
    chart.title.tx.rich.p[0].pPr.defRPr.b  = True

    chart.legend.position = "b"
    chart.legend.overlay = False
    chart.x_axis.title = None
    chart.y_axis.title = None
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.y_axis.number_format = "0%"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1


    # Text sizing/rotation
    AXIS_SZ = 1400
    LEGEND_SZ = 1300
    rot_units = -5400000 # -90° * 60000
    chart.x_axis.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=AXIS_SZ)),
                     endParaRPr=CharacterProperties())],
        bodyPr=RichTextProperties(rot=rot_units)
    )
    chart.y_axis.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=AXIS_SZ)),
                     endParaRPr=CharacterProperties())],
        bodyPr=RichTextProperties()
    )
    chart.legend.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=LEGEND_SZ)),
                     endParaRPr=CharacterProperties())],
        bodyPr=RichTextProperties()
    )

    # Add % series, color consistently
    for j in pct_cols:
        series_ref = Reference(ws7, min_col=j, min_row=1, max_col=j, max_row=last_row)
        chart.add_data(series_ref, titles_from_data=True)

        header = str(ws7.cell(row=1, column=j).value or "")
        hex6 = AREA_COLORS_HEX.get(header)
        if hex6:
            try:
                s = chart.series[-1]
                s.graphicalProperties.solidFill = hex6
            except Exception:
                pass
                
    # Categories FIRST (A2:A<last_row>)
    cats = Reference(ws7, min_col=1, min_row=2, max_col=1, max_row=last_row)
    chart.set_categories(cats)
    # Keep bars centered over ticks; add slight side padding via layout if you want
    chart.y_axis.crossBetween = "between"

    chart.height = 18
    chart.width  = 32
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    ws9.add_chart(chart, "B3")
    return wb
