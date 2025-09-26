import math
from typing import Optional, Dict, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def _norm(s: str) -> str:
    return str(s).strip().lower().replace(" ", "").replace("_", "")

def _resolve_cols(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    norm_to_orig = {_norm(c): c for c in df.columns}
    def pick(*aliases: str) -> Optional[str]:
        for a in aliases:
            k = _norm(a)
            if k in norm_to_orig:
                return norm_to_orig[k]
        return None
    return {
        "pixel_key": pick("Pixel_ID", "pixelid", "pixel"),
        "year": pick("year"),
        "payout_amount": pick("PayoutAmountBase", "payoutamountbase", "payout_amount", "payoutamount"),
        "payout_percent": pick("PayoutsPercent", "payoutspercent", "percent_payout"),
        "loan_amount": pick("Loan_Amount", "loanamount"),
        "sum_insured": pick("Sum_Insured", "suminsured"),
        "area": pick("area"),
        "region": pick("region"),
        "pixel_lon": pick("lon", "longitude"),
        "pixel_lat": pick("lat", "latitude"),
        "pixel_id": pick("Pixel_ID", "pixelid"),
    }

def _first_non_null(s: pd.Series):
    s2 = s.dropna()
    return s2.iloc[0] if not s2.empty else None

def autosize_columns(ws, col_start: int, col_end: int, min_width: int = 8, max_width: int = 45):
    for col in range(col_start, col_end + 1):
        m = 0
        for vals in ws.iter_cols(min_col=col, max_col=col, min_row=1, max_row=ws.max_row, values_only=True):
            for v in vals:
                if v is None:
                    continue
                m = max(m, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, min(max_width, m + 2))

def ws_title_is_default(title: str) -> bool:
    return str(title).lower().startswith("sheet")

def build_excel_sheet4(
    df: pd.DataFrame,
    wb: Optional[Workbook] = None,
    sheet_name: str = "4. Payout Statistics"
) -> Workbook:
    """
    Build a summary statistics sheet (no yearly rows) using pre‑computed payout
    columns:
        PayoutAvg, PayoutSD, PayoutMin, PayoutMax, Payout90, Payout95, PayoutCoV
    """
    cols = _resolve_cols(df)

    required_meta = ["pixel_key", "loan_amount", "sum_insured", "area", "region", "pixel_lon", "pixel_lat", "pixel_id"]
    for k in required_meta:
        if not cols[k]:
            # Not all strictly required (loan_amount / sum_insured may be None) but keep a notice.
            pass

    needed_stats = {
        "avg": "PayoutAvg",
        "sd": "PayoutSD",
        "min": "PayoutMin",
        "max": "PayoutMax",
        "p90": "Payout90",
        "p95": "Payout95",
        "cov": "PayoutCoV",
    }
    for alias, coln in needed_stats.items():
        if coln not in df.columns:
            raise ValueError(f"Missing required payout statistic column '{coln}' in dataframe.")

    pixel_col = cols["pixel_key"]

    # Keep one row per pixel (assume already aggregated). If duplicates, take first.
    df_pix = df.drop_duplicates(subset=[pixel_col]).copy()

    pixel_order = [p for p in df_pix[pixel_col] if pd.notnull(p)]

    # Metadata
    meta = {}
    for pix in pixel_order:
        sub = df_pix[df_pix[pixel_col] == pix]
        row = sub.iloc[0]
        meta[pix] = {
            "loan": _safe_get(row, cols["loan_amount"]),
            "sum_insured": _safe_get(row, cols["sum_insured"]),
            "area": _safe_get(row, cols["area"]),
            "region": _safe_get(row, cols["region"]),
            "lon": _safe_get(row, cols["pixel_lon"]),
            "lat": _safe_get(row, cols["pixel_lat"]),
            "pixelid": _safe_get(row, cols["pixel_id"]) or pix,
            "avg": _safe_get(row, needed_stats["avg"]),
            "sd": _safe_get(row, needed_stats["sd"]),
            "min": _safe_get(row, needed_stats["min"]),
            "max": _safe_get(row, needed_stats["max"]),
            "p90": _safe_get(row, needed_stats["p90"]),
            "p95": _safe_get(row, needed_stats["p95"]),
            "cov": _safe_get(row, needed_stats["cov"]),
        }

    total_pixels = len(pixel_order)
    total_loan = sum([m["loan"] for m in meta.values() if m["loan"] is not None])
    # Aggregate "Total" = sum of average payouts (can adjust if you want blank)
    total_payout_avg_sum = sum([m["avg"] for m in meta.values() if m["avg"] is not None])

    wb = wb or Workbook()
    if wb.active and ws_title_is_default(wb.active.title) and wb.active.max_row == 1:
        ws = wb.active
    else:
        ws = wb.create_sheet()
    ws.title = sheet_name

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    left = Alignment(horizontal="left")

    # Freeze (similar style as sheet 3): keep first 8 metadata rows (Pixel ID row is row 8), freeze at row 9 + after col E
    # Pixel ID will remain visible; columns A–E frozen
    ws.freeze_panes = ws.cell(row=9, column=2)  # F9

    area_colors_hex = {
        "Northern Zone": "1F77B4",
        "Central Zone": "2CA02C",
        "Lake Zone": "FF7F0E",
        "Western Zone": "9467BD",
        "Southern Highlands Zone": "8C564B",
        "Coastal Zone": "17BECF",
        "Zanzibar (Islands)": "7F7F7F",
    }
    def to_fill(hex6: str) -> PatternFill:
        return PatternFill(fill_type="solid", start_color=f"FF{hex6}", end_color=f"FF{hex6}")

    col_label = 1  # Column A will hold labels (simpler for a purely wide summary)
    first_data_col = 2  # Start pixel values at column B

    # Row mapping (aligning with provided CSV example):
    # 1 Pixel count
    # 2 Loan amounts (USD)
    # 3 Sum insured
    # 4 Area
    # 5 Region
    # 6 Pixel Lon
    # 7 Pixel Lat
    # 8 Pixel ID
    # 9 Total
    # 10 Average Payout by pixel
    # 11 SD
    # 12 CoV
    # 13 Min
    # 14 Max
    # 15 90th percentile
    # 16 95th percentile
    # (17 blank)

    # Row 1: Pixel count enumerator
    ws.cell(row=1, column=col_label).value = "Pixel count"
    for j, _ in enumerate(pixel_order):
        c = ws.cell(row=1, column=first_data_col + j)
        c.value = j + 1
        c.alignment = center

    # Row 2: Loan amounts
    ws.cell(row=2, column=col_label).value = "Loan amounts (USD)"
    for j, pix in enumerate(pixel_order):
        v = meta[pix]["loan"]
        c = ws.cell(row=2, column=first_data_col + j)
        c.value = v
        c.number_format = "$#,##0"

    # Row 3: Sum insured
    ws.cell(row=3, column=col_label).value = "Sum insured"
    for j, pix in enumerate(pixel_order):
        v = meta[pix]["sum_insured"]
        c = ws.cell(row=3, column=first_data_col + j)
        c.value = v
        c.number_format = "#,##0"

    # Row 4: Area
    ws.cell(row=4, column=col_label).value = "Area"
    for j, pix in enumerate(pixel_order):
        v = meta[pix]["area"]
        cell = ws.cell(row=4, column=first_data_col + j)
        cell.value = v
        if v:
            for name, hexv in area_colors_hex.items():
                if str(v).lower() == name.lower():
                    cell.fill = to_fill(hexv)
                    break

    # Row 5: Region
    ws.cell(row=5, column=col_label).value = "Region"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=5, column=first_data_col + j).value = meta[pix]["region"]

    # Row 6: Pixel Lon
    ws.cell(row=6, column=col_label).value = "Pixel Lon"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=6, column=first_data_col + j).value = meta[pix]["lon"]

    # Row 7: Pixel Lat
    ws.cell(row=7, column=col_label).value = "Pixel Lat"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=7, column=first_data_col + j).value = meta[pix]["lat"]

    # Row 8: Pixel ID
    ws.cell(row=8, column=col_label).value = "Pixel ID"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=8, column=first_data_col + j).value = meta[pix]["pixelid"]

    # Row 9: Total (just a simple label, no values)
    ws.cell(row=9, column=col_label).value = "Total"
    total_label_cell = ws.cell(row=9, column=col_label)
    total_label_cell.font = bold  # ensure bold explicitly
 
    # Rest of rows: payout statistics
    # Helper writer for metric rows
    def write_metric(row_idx: int, label: str, key: str, number_format="#,##0"):
        ws.cell(row=row_idx, column=col_label).value = label
        for j, pix in enumerate(pixel_order):
            v = meta[pix][key]
            c = ws.cell(row=row_idx, column=first_data_col + j)
            if pd.notnull(v):
                c.value = float(v)
            else:
                c.value = None
            c.number_format = number_format

    # Metrics
    write_metric(10, "Average Payout by pixel", "avg")
    write_metric(11, "SD", "sd")
    write_metric(12, "CoV", "cov")  # Coefficient of Variation (keep raw or format as number)
    write_metric(13, "Min", "min")
    write_metric(14, "Max", "max")
    write_metric(15, "90th percentile", "p90")
    write_metric(16, "95th percentile", "p95")

    # Bold label rows
    for r in range(1, 17):
        cell = ws.cell(row=r, column=col_label)
        if cell.value:
            cell.font = bold
            cell.alignment = left

    # Additional bold for total-like lines
    for r in (2, 3, 9, 10, 11):
        ws.cell(row=r, column=col_label).font = bold

    # Autosize columns
    autosize_columns(ws, 1, first_data_col + len(pixel_order) - 1)


    return wb

def _safe_get(row, colname):
    if not colname:
        return None
    try:
        return row[colname]
    except Exception:
        return None

# Example usage:
# wb = build_payout_statistics_sheet(df_summary)
# wb.save("policy_pilot_payout_statistics.xlsx")
 