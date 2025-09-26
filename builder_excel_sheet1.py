# %%
import math
from typing import Optional, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def to_fill(hex6: str) -> PatternFill:
    return PatternFill(fill_type="solid", start_color=f"FF{hex6}", end_color=f"FF{hex6}")

def _norm(s: str) -> str:
    return str(s).strip().lower().replace(" ", "").replace("_", "")
def _resolve_cols(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    norm_to_orig = {_norm(c): c for c in df.columns}
    def pick(*aliases: str) -> Optional[str]:
        for a in aliases:
            key = _norm(a)
            if key in norm_to_orig:
                return norm_to_orig[key]
        return None
    return {
        # Keep Pixel_ID as the pivot ('pixel') key; do NOT fall back to Index_ID
        "pixel_key": pick("Pixel_ID", "pixel"),
        "year": pick("year"),
        "yield": pick("Yield_Abs"),
        "area": pick("area"),
        "region": pick("region"),
        "pixel_lon": pick("lon", "longitude", "pixel lon"),
        "pixel_lat": pick("lat", "latitude", "pixel lat"),
        "pixel_id": pick("Pixel_ID", "pixelid"),
        "index_id": pick("Index_ID", "indexid", "index"),
    }

def _first_non_null(s: pd.Series):
    s2 = s.dropna()
    return s2.iloc[0] if not s2.empty else None

def autosize_columns(ws, col_start: int, col_end: int, min_width: int = 8, max_width: int = 40):
    for col in range(col_start, col_end + 1):
        max_len = 0
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=1, max_row=ws.max_row, values_only=True):
            for v in cell:
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, min(max_width, max_len + 2))

def ws_title_is_default(title: str) -> bool:
    return str(title).lower().startswith("sheet")

def build_excel_sheet1(
    df: pd.DataFrame,
    wb: Optional[Workbook] = None,
    sheet_name: str = "1. Modelled Yield"
) -> Workbook:
    cols = _resolve_cols(df)
    required = {k: cols[k] for k in ("pixel_key", "year", "yield")}
    missing = [k for k, v in required.items() if not v]
    if missing:
        raise ValueError(f"Missing required columns: {missing}. Available: {list(df.columns)}")

    pixel_col = cols["pixel_key"]
    year_col = cols["year"]
    yield_col = cols["yield"]

    # Pivot first and derive pixel order from pivot columns to preserve dtype
    pivot = df.pivot_table(index=year_col, columns=pixel_col, values=yield_col, aggfunc="first").sort_index()
    pixel_order: List = [c for c in list(pivot.columns) if not pd.isna(c)]

    # Per-pixel metadata
    meta: Dict[object, Dict[str, Optional[object]]] = {}
    for pix in pixel_order:
        sub = df[df[pixel_col] == pix]
        lon = _first_non_null(sub[cols["pixel_lon"]]) if cols["pixel_lon"] else None
        lat = _first_non_null(sub[cols["pixel_lat"]]) if cols["pixel_lat"] else None
        meta[pix] = {
            "area": _first_non_null(sub[cols["area"]]) if cols["area"] else None,
            "region": _first_non_null(sub[cols["region"]]) if cols["region"] else None,
            # Use the real Index_ID column here
            "indexid": _first_non_null(sub[cols["index_id"]]) if cols.get("index_id") else None,
            "lon": float(lon) if lon is not None and pd.notna(lon) else None,
            "lat": float(lat) if lat is not None and pd.notna(lat) else None,
            "pixelid": _first_non_null(sub[cols["pixel_id"]]) if cols["pixel_id"] else None,
        }

    # Workbook/sheet
    wb = wb or Workbook()
    ws = wb.active if (wb.active and wb.active.max_row == 1 and ws_title_is_default(wb.active.title)) else wb.create_sheet()
    ws.title = sheet_name
    # Freeze top rows and left columns
    ws.freeze_panes = ws.cell(row=10, column=6)  # F10

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    left = Alignment(horizontal="left")

    col_label = 5          # E
    first_data_col = 6     # F
    row_title = 2
    row_meta_start = 3
    row_data_start = 10

    # Title
    ws.cell(row=row_title, column=col_label).value = "MODELLED YIELDS (tons per ha)"
    ws.cell(row=row_title, column=col_label).font = bold
    ws.cell(row=row_title, column=col_label).alignment = left
    #Fill title cell with yellow
    hx = "FFFF00"
    cell = ws.cell(row=row_title, column=col_label)
    cell.fill = to_fill(hx)

    # Pixel count
    ws.cell(row=row_meta_start + 0, column=col_label).value = "Pixel count"
    for j, _pix in enumerate(pixel_order):
        c = ws.cell(row=row_meta_start + 0, column=first_data_col + j)
        c.value = j + 1
        c.alignment = center

    # Area
    ws.cell(row=row_meta_start + 1, column=col_label).value = "Area"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=row_meta_start + 1, column=first_data_col + j).value = meta[pix]["area"] or ""

    # Region
    ws.cell(row=row_meta_start + 2, column=col_label).value = "Region"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=row_meta_start + 2, column=first_data_col + j).value = meta[pix]["region"] or ""

    # Index ID
    ws.cell(row=row_meta_start + 3, column=col_label).value = "Index ID"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=row_meta_start + 3, column=first_data_col + j).value = meta[pix]["indexid"] or ""

    # Pixel Lon
    ws.cell(row=row_meta_start + 4, column=col_label).value = "Pixel Lon"
    for j, pix in enumerate(pixel_order):
        v = meta[pix]["lon"]
        ws.cell(row=row_meta_start + 4, column=first_data_col + j).value = float(v) if isinstance(v, (int, float)) and not math.isnan(v) else v

    # Note + Pixel Lat
    ws.cell(row=row_meta_start + 5, column=1).value = "Note: if yield data absent, then pixel has dropped out"
    ws.cell(row=row_meta_start + 5, column=col_label).value = "Pixel Lat"
    for j, pix in enumerate(pixel_order):
        v = meta[pix]["lat"]
        ws.cell(row=row_meta_start + 5, column=first_data_col + j).value = float(v) if isinstance(v, (int, float)) and not math.isnan(v) else v

    # Pixel ID
    ws.cell(row=row_meta_start + 6, column=col_label).value = "Pixel ID"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=row_meta_start + 6, column=first_data_col + j).value = meta[pix]["pixelid"] or pix

    # Data rows: use positional indexing to avoid dtype mismatches
    years = list(pivot.index)
    for i, y in enumerate(years, start=1):
        r = row_data_start + i - 1
        # Index
        c_idx = ws.cell(row=r, column=col_label - 1)
        c_idx.value = i
        c_idx.alignment = center
        # Year
        try:
            ws.cell(row=r, column=col_label).value = int(y)
        except Exception:
            ws.cell(row=r, column=col_label).value = y
        # Yields
        for j in enumerate(pixel_order):
            pass  # placeholder

        for j in range(len(pixel_order)):
            val = pivot.iat[i - 1, j]
            ws.cell(row=r, column=first_data_col + j).value = (float(val) if pd.notnull(val) else None)

    # Bold labels in column E
    for rr in range(row_meta_start, row_meta_start + 7):
        c = ws.cell(row=rr, column=col_label)
        c.font = bold
        c.alignment = left

    last_col = max(first_data_col + len(pixel_order) - 1, col_label)
    autosize_columns(ws, 1, last_col)

    return wb

