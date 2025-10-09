import pandas as pd
from typing import Optional, Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

# --- constants for cross-sheet references ---
SHEET1_NAME = "1. Modelled Yield"
SHEET1_ROW_FARMERCOUNT = 6   # row of "Farmer count" in Sheet 1
COL_FIRST_PIXEL = 6          # F
COL_YEAR_LABEL  = 5          # E

def _norm(s: str) -> str:
    return str(s).strip().lower().replace(" ", "").replace("_", "")

def _resolve_cols(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    norm_to_orig = {_norm(c): c for c in df.columns}
    def pick(*aliases: str) -> Optional[str]:
        for a in aliases:
            k = _norm(a)
            if k in norm_to_orig:
                return norm_to_orig[k]
        return None
    return {
        "pixel_key": pick("Pixel_ID", "pixelid", "pixel"),
        "year":      pick("Year", "year"),
        "loan":      pick("Pixel_Loan_Amount"),
        "area":      pick("Area", "area_ha", "hectares"),
        "region":    pick("Region"),
        "lon":       pick("lon", "longitude"),
        "lat":       pick("lat", "latitude"),
        "pixel_id":  pick("Pixel_ID", "pixelid"),
    }

def _first_non_null(s: pd.Series):
    s2 = s.dropna()
    return s2.iloc[0] if not s2.empty else None

def _fill(hex6: str) -> PatternFill:
    return PatternFill(fill_type="solid", start_color=f"FF{hex6}", end_color=f"FF{hex6}")

def _autosize(ws, c1: int, c2: int, min_w: int = 8, max_w: int = 40):
    for col in range(c1, c2 + 1):
        m = 0
        for vals in ws.iter_cols(min_col=col, max_col=col, min_row=1, max_row=ws.max_row, values_only=True):
            for v in vals:
                if v is None:
                    continue
                m = max(m, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = max(min_w, min(max_w, m + 2))

def _to_float_or_none(x):
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return None
        return float(x)
    except Exception:
        return None

def ws_title_is_default(title: str) -> bool:
    return str(title).lower().startswith("sheet")

def build_excel_sheet3(
    df: pd.DataFrame,
    wb: Optional[Workbook] = None,
    sheet_name: str = "3. Payout Amounts"
) -> Workbook:
    """
    Sheet 3 (Payout Amounts):
    - Loan Amounts (USD) is now a FORMULA: (Sheet1 Farmer count) × (per-farmer region loan from data).
    - Sum Insured = 0.4 × Loan (formula, blank-safe).
    - Grid uses '2. Payouts %' × Sum Insured (blank-safe).
    - Stats block identical to before.
    """
    cols = _resolve_cols(df)
    if not cols["pixel_key"] or not cols["year"]:
        raise ValueError("Dataframe must include Pixel_ID and Year.")

    pixel_col = cols["pixel_key"]; year_col = cols["year"]
    pixel_order = sorted(df[pixel_col].dropna().unique().tolist())
    year_list   = sorted(df[year_col].dropna().unique().tolist())

    # Per-pixel metadata (NO farmer count added here)
    meta: Dict[object, Dict[str, Optional[object]]] = {}
    for pix in pixel_order:
        sub = df[df[pixel_col] == pix]
        meta[pix] = {
            "loan":    _first_non_null(sub[cols["loan"]])    if cols["loan"]   else None,  # per-farmer regional loan
            "area":    _first_non_null(sub[cols["area"]])    if cols["area"]   else None,
            "region":  _first_non_null(sub[cols["region"]])  if cols["region"] else None,
            "lon":     _first_non_null(sub[cols["lon"]])     if cols["lon"]    else None,
            "lat":     _first_non_null(sub[cols["lat"]])     if cols["lat"]    else None,
            "pixelid": _first_non_null(sub[cols["pixel_id"]]) if cols["pixel_id"] else pix,
        }

    wb = wb or Workbook()
    ws = wb.active if (wb.active and wb.active.max_row == 1 and ws_title_is_default(wb.active.title)) else wb.create_sheet()
    ws.title = sheet_name

    bold = Font(bold=True)
    red_font = Font(color="FF0000")
    left = Alignment(horizontal="left")
    center = Alignment(horizontal="center")

    # Layout constants
    ROW_TITLE       = 1
    ROW_META_START  = 2   # rows 2..8 are metadata
    ROW_PIXEL_ID    = 9   # header row
    ROW_FIRST_DATA  = 10  # first data row
    ws.freeze_panes = ws.cell(row=ROW_FIRST_DATA, column=COL_FIRST_PIXEL)

    # Title
    ws.cell(row=ROW_TITLE, column=COL_YEAR_LABEL).value = "PAYOUT AMOUNTS (USD)"
    ws.cell(row=ROW_TITLE, column=COL_YEAR_LABEL).font = bold
    ws.cell(row=ROW_TITLE, column=COL_YEAR_LABEL).alignment = left
    ws.cell(row=ROW_TITLE, column=COL_YEAR_LABEL).fill = _fill("FFFF00")

    # ===== Summary (unchanged) =====
    ws.cell(row=3, column=1).value = "Total Loan Amounts (USD)"; ws.cell(row=3, column=1).font = bold
    ws.cell(row=5, column=1).value = "Total Number of Pixels";   ws.cell(row=5, column=1).font = bold

    last_col_idx = COL_FIRST_PIXEL + len(pixel_order) - 1
    first_col_letter = get_column_letter(COL_FIRST_PIXEL)
    last_col_letter  = get_column_letter(last_col_idx)
    loans_row = ROW_META_START + 1  # Loan Amounts row index
    loans_rng = f"{first_col_letter}{loans_row}:{last_col_letter}{loans_row}"

    ws.cell(row=3, column=3).value = f"=IF(COUNT({loans_rng})=0,\"\",SUM({loans_rng}))"
    ws.cell(row=3, column=3).number_format = "#,##0"
    loan_row_range = f"${first_col_letter}${loans_row}:${last_col_letter}${loans_row}"
    ws.cell(row=5, column=3).value = f"=COUNT({loan_row_range})"

    # Note in D1:D7 (merged, red)
    ws.merge_cells(start_row=1, start_column=4, end_row=7, end_column=4)
    note_cell = ws.cell(row=1, column=4)
    note_cell.value = ("Note: Loan amounts are given as average loan amount across pixels within a given region, "
                       "as pixel-level loan amount info not available yet.")
    note_cell.font = red_font
    note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ===== Metadata labels (E), values F→ =====
    ws.cell(row=ROW_META_START + 0, column=COL_YEAR_LABEL).value = "Pixel count"
    ws.cell(row=ROW_META_START + 1, column=COL_YEAR_LABEL).value = "Loan Amounts (USD)"
    ws.cell(row=ROW_META_START + 2, column=COL_YEAR_LABEL).value = "Sum Insured (USD)"
    ws.cell(row=ROW_META_START + 3, column=COL_YEAR_LABEL).value = "Area"
    ws.cell(row=ROW_META_START + 4, column=COL_YEAR_LABEL).value = "Region"
    ws.cell(row=ROW_META_START + 5, column=COL_YEAR_LABEL).value = "Pixel Lon"
    ws.cell(row=ROW_META_START + 6, column=COL_YEAR_LABEL).value = "Pixel Lat"

    # Area colors
    area_colors_hex = {
        "Northern Zone": "1F77B4", "Central Zone": "2CA02C", "Lake Zone": "FF7F0E",
        "Western Zone": "9467BD", "Southern Highlands Zone": "8C564B",
        "Coastal Zone": "17BECF", "Zanzibar (Islands)": "7F7F7F",
    }

    # === Write metadata rows ===
    for j, pix in enumerate(pixel_order):
        c = COL_FIRST_PIXEL + j
        colL = get_column_letter(c)

        # Pixel count ordinal
        ws.cell(row=ROW_META_START + 0, column=c).value = j + 1

        # --- Update: Loan Amounts (USD) = Pixel-total loan---
        loan_pixel = _to_float_or_none(meta[pix]["loan"])
        if loan_pixel is not None:
            farmer_ref = f"'{SHEET1_NAME}'!{colL}{SHEET1_ROW_FARMERCOUNT}"
            loan_formula = (
                f"=IF(OR({farmer_ref}=\"\",NOT(ISNUMBER({farmer_ref}))),\"\",{loan_pixel})"
            )
            ws.cell(row=ROW_META_START + 1, column=c).value = loan_formula
            ws.cell(row=ROW_META_START + 1, column=c).number_format = "#,##0"
        else:
            ws.cell(row=ROW_META_START + 1, column=c).value = ""
            ws.cell(row=ROW_META_START + 1, column=c).number_format = "#,##0"

        # Sum Insured (formula = 0.4*Loan, blank-safe)
        loan_ref = f"{colL}{ROW_META_START + 1}"
        ws.cell(row=ROW_META_START + 2, column=c).value = f"=IF({loan_ref}=\"\",\"\",0.4*{loan_ref})"
        ws.cell(row=ROW_META_START + 2, column=c).number_format = "#,##0"

        # Area (colored if matched)
        area_cell = ws.cell(row=ROW_META_START + 3, column=c)
        v_area = meta[pix]["area"] or ""
        area_cell.value = v_area
        hx = area_colors_hex.get(str(v_area))
        if not hx and v_area:
            for k, val in area_colors_hex.items():
                if str(v_area).strip().lower() == k.lower():
                    hx = val; break
        if hx:
            area_cell.fill = _fill(hx)

        ws.cell(row=ROW_META_START + 4, column=c).value = meta[pix]["region"] or ""
        ws.cell(row=ROW_META_START + 5, column=c).value = meta[pix]["lon"]
        ws.cell(row=ROW_META_START + 6, column=c).value = meta[pix]["lat"]

    # ===== Header row (row 9) =====
    ws.cell(row=ROW_PIXEL_ID, column=2).value = "SD"
    ws.cell(row=ROW_PIXEL_ID, column=4).value = "Average"
    ws.cell(row=ROW_PIXEL_ID, column=COL_YEAR_LABEL).value = "Pixel ID"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=ROW_PIXEL_ID, column=COL_FIRST_PIXEL + j).value = meta[pix]["pixelid"] or str(pix)

    # Year labels (E10..)
    for i, y in enumerate(year_list):
        r = ROW_FIRST_DATA + i
        try:
            ws.cell(row=r, column=COL_YEAR_LABEL).value = int(y)
        except Exception:
            ws.cell(row=r, column=COL_YEAR_LABEL).value = y

    # ===== GRID: payout amount = '2. Payouts %' × SumInsured (blank-safe) =====
    sheet2_name = "2. Payouts %"
    last_col_idx = COL_FIRST_PIXEL + len(pixel_order) - 1
    last_col_letter = get_column_letter(last_col_idx)

    for i, _ in enumerate(year_list):
        r = ROW_FIRST_DATA + i
        row_rng = f"{get_column_letter(COL_FIRST_PIXEL)}{r}:{last_col_letter}{r}"

        # Per-year stats (A–D), blank-safe (AMOUNTS)
        ws.cell(row=r, column=2).value = f"=IF(COUNT({row_rng})<=1,\"\",STDEV({row_rng}))"
        ws.cell(row=r, column=4).value = f"=IF(COUNT({row_rng})=0,\"\",AVERAGE({row_rng}))"
        ws.cell(row=r, column=2).number_format = "#,##0"
        ws.cell(row=r, column=4).number_format = "#,##0"

        # Grid cells
        for j, _ in enumerate(pixel_order):
            c = COL_FIRST_PIXEL + j
            colL = get_column_letter(c)
            pct_ref = f"'{sheet2_name}'!{colL}{r}"
            si_ref  = f"{colL}{ROW_META_START + 2}"  # Sum Insured row
            formula = (
                f"=IF(OR(ISBLANK({pct_ref}),NOT(ISNUMBER({pct_ref}))),\"\","
                f"IF(OR(ISBLANK({si_ref}),NOT(ISNUMBER({si_ref}))),\"\",{pct_ref}*{si_ref}))"
            )
            cell = ws.cell(row=r, column=c)
            cell.value = formula
            cell.number_format = "#,##0"

    # ===== Statistics block =====
    end_row = ROW_FIRST_DATA + len(year_list) - 1
    r_sum1 = end_row + 1

    # Average SD (B)
    ws.cell(row=r_sum1, column=1).value = "Average SD"; ws.cell(row=r_sum1, column=1).font = bold
    ws.cell(row=r_sum1, column=2).value = (
        f"=IF(COUNT(B{ROW_FIRST_DATA}:B{end_row})=0,\"\",AVERAGE(B{ROW_FIRST_DATA}:B{end_row}))"
    )
    ws.cell(row=r_sum1, column=2).number_format = "#,##0"

    # Average payout (per pixel per year) in C→D
    ws.cell(row=r_sum1, column=3).value = "Average payout (per pixel per year)"
    ws.cell(row=r_sum1, column=3).font = bold
    grid_rng = f"{get_column_letter(COL_FIRST_PIXEL)}{ROW_FIRST_DATA}:{get_column_letter(last_col_idx)}{end_row}"
    ws.cell(row=r_sum1, column=4).value = f"=IF(COUNT({grid_rng})=0,\"\",AVERAGE({grid_rng}))"
    ws.cell(row=r_sum1, column=4).number_format = "#,##0"

    # Average Payout by pixel (E + across pixel columns)
    ws.cell(row=r_sum1, column=COL_YEAR_LABEL).value = "Average Payout by pixel"
    ws.cell(row=r_sum1, column=COL_YEAR_LABEL).font = Font(bold=True)
    for j in range(len(pixel_order)):
        c = COL_FIRST_PIXEL + j
        colL = get_column_letter(c)
        col_rng = f"{colL}{ROW_FIRST_DATA}:{colL}{end_row}"
        ws.cell(row=r_sum1, column=c).value = f"=IF(COUNT({col_rng})=0,\"\",AVERAGE({col_rng}))"
        ws.cell(row=r_sum1, column=c).number_format = "#,##0"

    # Overall SD (B) and per-pixel SD row (E + across pixel columns)
    r_sd = r_sum1 + 1
    ws.cell(row=r_sd, column=1).value = "Overall SD"; ws.cell(row=r_sd, column=1).font = bold
    ws.cell(row=r_sd, column=2).value = f"=IF(COUNT(D{ROW_FIRST_DATA}:D{end_row})<=1,\"\",STDEV(D{ROW_FIRST_DATA}:D{end_row}))"
    ws.cell(row=r_sd, column=2).number_format = "#,##0"
    ws.cell(row=r_sd, column=COL_YEAR_LABEL).value = "SD"; ws.cell(row=r_sd, column=COL_YEAR_LABEL).font = bold
    for j in range(len(pixel_order)):
        c = COL_FIRST_PIXEL + j
        colL = get_column_letter(c)
        col_rng = f"{colL}{ROW_FIRST_DATA}:{colL}{end_row}"
        ws.cell(row=r_sd, column=c).value = f"=IF(COUNT({col_rng})<=1,\"\",STDEV({col_rng}))"
        ws.cell(row=r_sd, column=c).number_format = "#,##0"

    # Min / Max / Percentiles (per-pixel)
    for offset, label, fbuild in [
        (1, "Min", lambda rng: f"=IF(COUNT({rng})=0,\"\",MIN({rng}))"),
        (2, "Max", lambda rng: f"=IF(COUNT({rng})=0,\"\",MAX({rng}))"),
        (3, "90th percentile", lambda rng: f"=IF(COUNT({rng})=0,\"\",PERCENTILE({rng},0.9))"),
        (4, "95th percentile", lambda rng: f"=IF(COUNT({rng})=0,\"\",PERCENTILE({rng},0.95))"),
    ]:
        r = r_sd + offset
        ws.cell(row=r, column=COL_YEAR_LABEL).value = label
        ws.cell(row=r, column=COL_YEAR_LABEL).font = bold
        for j in range(len(pixel_order)):
            c = COL_FIRST_PIXEL + j
            colL = get_column_letter(c)
            col_rng = f"{colL}{ROW_FIRST_DATA}:{colL}{end_row}"
            ws.cell(row=r, column=c).value = fbuild(col_rng)
            ws.cell(row=r, column=c).number_format = "#,##0"

    # ===== Styling & sizing =====
    for rr in range(ROW_META_START, ROW_PIXEL_ID + 1):
        ws.cell(row=rr, column=COL_YEAR_LABEL).font = bold
        ws.cell(row=rr, column=COL_YEAR_LABEL).alignment = left
    for cc in [2, 4]:
        ws.cell(row=ROW_PIXEL_ID, column=cc).font = bold
        ws.cell(row=ROW_PIXEL_ID, column=cc).alignment = center
    for j in range(len(pixel_order)):
        ws.cell(row=ROW_PIXEL_ID, column=COL_FIRST_PIXEL + j).font = bold
        ws.cell(row=ROW_PIXEL_ID, column=COL_FIRST_PIXEL + j).alignment = center

    last_col = max(COL_FIRST_PIXEL + len(pixel_order) - 1, COL_YEAR_LABEL)
    _autosize(ws, 1, last_col)

    # Force full recalc on open
    try:
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    except Exception:
        try:
            wb.calculation.fullCalcOnLoad = True
        except Exception:
            pass



    # === BEGIN: Formatting tweaks per request (v2) ===

    from openpyxl.styles import Alignment as _Align3

    from openpyxl.utils import get_column_letter as _gcl3


    # Task 3: Merge 'Total Loan Amounts (USD)' (A3) with B3 and 'Total Number of Pixels' (A5) with B5

    try:

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)

        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)

    except Exception:

        pass


    # Task 4: Merge title with neighbor on the right; set column E width -> 18.0

    try:

        ws.merge_cells(start_row=ROW_TITLE, start_column=COL_YEAR_LABEL, end_row=ROW_TITLE, end_column=COL_FIRST_PIXEL)

    except Exception:

        pass

    ws.column_dimensions['E'].width = 18.0


    # Task 2: Columns A-D width -> 18.0

    for _col in ['A','B','C','D']:

        ws.column_dimensions[_col].width = 18.0


    # Task 2: Row with 'Average SD' -> set height 26.7 and wrap text

    try:

        try:

            last_col_for_wrap = max(COL_FIRST_PIXEL + len(pixel_order) - 1, COL_YEAR_LABEL)

        except Exception:

            last_col_for_wrap = ws.max_column

        for _cc in range(1, last_col_for_wrap + 1):

            _cell = ws.cell(row=r_sum1, column=_cc)

            _cell.alignment = _Align3(wrap_text=True, horizontal=_cell.alignment.horizontal if _cell.alignment else None, vertical=_cell.alignment.vertical if _cell.alignment else None)

        ws.row_dimensions[r_sum1].height = 26.7

    except Exception:

        pass


    # Task 1: Set data columns (F and onward) width -> 21.4

    _last_data_col = COL_FIRST_PIXEL + len(pixel_order) - 1

    for _c in range(COL_FIRST_PIXEL, max(COL_FIRST_PIXEL, _last_data_col) + 1):

        ws.column_dimensions[_gcl3(_c)].width = 21.4

    # === END: Formatting tweaks per request (v2) ===


    return wb
