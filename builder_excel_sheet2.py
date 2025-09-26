import math
from typing import Optional, Dict, List
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def _norm(s: str) -> str:
    return str(s).strip().lower().replace(" ", "").replace("_", "")

def _resolve_cols(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    norm_to_orig = {_norm(c): c for c in df.columns}
    def pick(*aliases: str) -> Optional[str]:
        for a in aliases:
            key = _norm(a)
            if key in norm_to_orig:
                return norm_to_orig[key]
        return None
    return {
        "pixel_key": pick("Pixel_ID", "pixelid", "pixel"),
        "year": pick("year"),
        "payout": pick("PayoutsPercent", "Percent_Payout", "payout_percent", "payouts_percent"),
        "attach": pick("Attach", "attach_threshold", "attach_kg_ha"),
        "detach": pick("Detach", "detach_threshold", "detach_kg_ha"),
        "area": pick("area"),
        "region": pick("region"),
        "pixel_lon": pick("lon", "longitude", "pixel lon"),
        "pixel_lat": pick("lat", "latitude", "pixel lat"),
        "pixel_id": pick("Pixel_ID", "pixelid"),
    }

def _first_non_null(s: pd.Series):
    s2 = s.dropna()
    return s2.iloc[0] if not s2.empty else None

def autosize_columns(ws, col_start: int, col_end: int, min_width: int = 8, max_width: int = 40):
    from openpyxl.utils import get_column_letter
    for col in range(col_start, col_end + 1):
        max_len = 0
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=1, max_row=ws.max_row, values_only=True):
            for v in cell:
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, min(max_width, max_len + 2))

def ws_title_is_default(title: str) -> bool:
    return str(title).lower().startswith("sheet")

def build_excel_sheet2(
    df: pd.DataFrame,
    wb: Optional[Workbook] = None,
    sheet_name: str = "2. Payouts %"
) -> Workbook:
    cols = _resolve_cols(df)
    required = {k: cols[k] for k in ("pixel_key", "year", "payout")}
    missing = [k for k, v in required.items() if not v]
    if missing:
        raise ValueError(f"Missing required columns: {missing}. Available: {list(df.columns)}")

    pixel_col = cols["pixel_key"]
    year_col = cols["year"]
    payout_col = cols["payout"]

    # Pivot payouts: rows=Year, cols=Pixel_ID
    pivot = df.pivot_table(index=year_col, columns=pixel_col, values=payout_col, aggfunc="first").sort_index()
    pixel_order: List = [c for c in list(pivot.columns) if not pd.isna(c)]

    # Per-pixel metadata (first non-null)
    meta = {}
    for pix in pixel_order:
        sub = df[df[pixel_col] == pix]
        meta[pix] = {
            "attach": _first_non_null(sub[cols["attach"]]) if cols["attach"] else None,
            "detach": _first_non_null(sub[cols["detach"]]) if cols["detach"] else None,
            "area": _first_non_null(sub[cols["area"]]) if cols["area"] else None,
            "region": _first_non_null(sub[cols["region"]]) if cols["region"] else None,
            "lon": _first_non_null(sub[cols["pixel_lon"]]) if cols["pixel_lon"] else None,
            "lat": _first_non_null(sub[cols["pixel_lat"]]) if cols["pixel_lat"] else None,
            "pixelid": _first_non_null(sub[cols["pixel_id"]]) if cols["pixel_id"] else None,
        }

    # Per-year stats
    year_avg = pivot.mean(axis=1, skipna=True)
    year_sd = pivot.std(axis=1, ddof=0, skipna=True)

    # Summary stats
    avg_sd = float(year_sd.mean(skipna=True)) if len(year_sd) else None
    avg_payout_overall = float(year_avg.mean(skipna=True)) if len(year_avg) else None
    by_pixel_avg = pivot.mean(axis=0, skipna=True)
    by_pixel_sd = pivot.std(axis=0, ddof=0, skipna=True)
    by_pixel_min = pivot.min(axis=0, skipna=True)
    by_pixel_max = pivot.max(axis=0, skipna=True)
    by_pixel_p90 = pivot.quantile(0.9, axis=0, interpolation="linear")
    by_pixel_p95 = pivot.quantile(0.95, axis=0, interpolation="linear")

    # Workbook/sheet

    wb = wb or Workbook()
    ws = wb.active if (wb.active and wb.active.max_row == 1 and ws_title_is_default(wb.active.title)) else wb.create_sheet()
    ws.title = sheet_name
    # Freeze first five columns (A–E) and top 10 rows (metadata)
    ws.freeze_panes = ws.cell(row=11, column=6)  # 'F11'

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    left = Alignment(horizontal="left")

    # Area color palette (no yellow)
    area_colors_hex = {
        "Northern Zone": "1F77B4",          # blue
        "Central Zone": "2CA02C",           # green
        "Lake Zone": "FF7F0E",              # orange
        "Western Zone": "9467BD",           # purple
        "Southern Highlands Zone": "8C564B",# brown
        "Coastal Zone": "17BECF",           # teal
        "Zanzibar (Islands)": "7F7F7F",     # gray
    }
    def to_fill(hex6: str) -> PatternFill:
        return PatternFill(fill_type="solid", start_color=f"FF{hex6}", end_color=f"FF{hex6}")


    # Layout similar to sheet1
    col_label = 5          # E
    first_data_col = 6     # F
    row_title = 2
    row_meta_start = 3
    row_header = row_meta_start + 7   # 10
    row_data_start = row_header + 1   # 11

    # Title
    ws.cell(row=row_title, column=col_label).value = "PAYOUTS % (fraction of sum insured)"
    ws.cell(row=row_title, column=col_label).font = bold
    ws.cell(row=row_title, column=col_label).alignment = left
    #Fill title cell with yellow
    hx = "FFFF00"
    cell = ws.cell(row=row_title, column=col_label)
    cell.fill = to_fill(hx)
    
    # Metadata rows
    # Row: Pixel count
    ws.cell(row=row_meta_start + 0, column=col_label).value = "Pixel count"
    for j, _pix in enumerate(pixel_order):
        c = ws.cell(row=row_meta_start + 0, column=first_data_col + j)
        c.value = j + 1
        c.alignment = center

    # Row: Attach
    ws.cell(row=row_meta_start + 1, column=col_label).value = "Attach (kg per ha)"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=row_meta_start + 1, column=first_data_col + j).value = meta[pix]["attach"]

    # Row: Detach
    ws.cell(row=row_meta_start + 2, column=col_label).value = "Detach (kg per ha)"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=row_meta_start + 2, column=first_data_col + j).value = meta[pix]["detach"]

    # Row: Area (write value and apply color fill)
    ws.cell(row=row_meta_start + 3, column=col_label).value = "Area"
    for j, pix in enumerate(pixel_order):
        cell = ws.cell(row=row_meta_start + 3, column=first_data_col + j)
        v = meta[pix]["area"] or ""
        cell.value = v
        hx = area_colors_hex.get(str(v), area_colors_hex.get(str(v).strip(), None))
        if not hx:
            # try case-insensitive match
            for k, val in area_colors_hex.items():
                if str(v).strip().lower() == k.lower():
                    hx = val
                    break
        if hx:
            cell.fill = to_fill(hx)

    # Row: Region
    ws.cell(row=row_meta_start + 4, column=col_label).value = "Region"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=row_meta_start + 4, column=first_data_col + j).value = meta[pix]["region"] or ""

    # Row: Pixel Lon
    ws.cell(row=row_meta_start + 5, column=col_label).value = "Pixel Lon"
    for j, pix in enumerate(pixel_order):
        v = meta[pix]["lon"]
        ws.cell(row=row_meta_start + 5, column=first_data_col + j).value = v

    # Row: Pixel Lat
    ws.cell(row=row_meta_start + 6, column=col_label).value = "Pixel Lat"
    for j, pix in enumerate(pixel_order):
        v = meta[pix]["lat"]
        ws.cell(row=row_meta_start + 6, column=first_data_col + j).value = v

    # Header row for data block
    ws.cell(row=row_header, column=2).value = "SD"
    ws.cell(row=row_header, column=4).value = "Average"
    ws.cell(row=row_header, column=col_label).value = "Pixel ID"
    for j, pix in enumerate(pixel_order):
        ws.cell(row=row_header, column=first_data_col + j).value = meta[pix]["pixelid"] or str(pix)

    # Data rows by year: SD | Average | Year | per-pixel payouts
    years = list(pivot.index)
    for i, y in enumerate(years, start=1):
        r = row_data_start + i - 1
        # Column B: SD for the year
        c_sd = ws.cell(row=r, column=2)
        val_sd = year_sd.iat[i - 1] if len(year_sd) >= i else None
        c_sd.value = float(val_sd) if pd.notnull(val_sd) else None
        c_sd.number_format = "0.00%"
        # Column D: Average for the year
        c_avg = ws.cell(row=r, column=4)
        val_avg = year_avg.iat[i - 1] if len(year_avg) >= i else None
        c_avg.value = float(val_avg) if pd.notnull(val_avg) else None
        c_avg.number_format = "0.00%"
        # Column E: Year
        try:
            ws.cell(row=r, column=col_label).value = int(y)
        except Exception:
            ws.cell(row=r, column=col_label).value = y
        # Columns F..: payouts per pixel
        for j in range(len(pixel_order)):
            v = pivot.iat[i - 1, j]
            c = ws.cell(row=r, column=first_data_col + j)
            c.value = float(v) if pd.notnull(v) else None
            c.number_format = "0.00%"


    # Summary rows (below data block)
    r_sum1 = row_data_start + len(years)
    ws.cell(row=r_sum1, column=1).value = "Average SD"
    c_avg_sd = ws.cell(row=r_sum1, column=2)
    c_avg_sd.value = avg_sd if avg_sd is not None and pd.notnull(avg_sd) else None
    c_avg_sd.number_format = "0.00%"
    ws.cell(row=r_sum1, column=1).font = bold

    ws.cell(row=r_sum1, column=3).value = "Average payout (% of sum assured)"
    c_avg_overall = ws.cell(row=r_sum1, column=4)
    c_avg_overall.value = avg_payout_overall if avg_payout_overall is not None and pd.notnull(avg_payout_overall) else None
    c_avg_overall.number_format = "0.00%"
    ws.cell(row=r_sum1, column=3).font = bold

    # Labels in column E
    ws.cell(row=r_sum1, column=col_label).value = "Average Payout by pixel"
    ws.cell(row=r_sum1, column=col_label).font = bold
    for j, pix in enumerate(pixel_order):
        c = ws.cell(row=r_sum1, column=first_data_col + j)
        v = by_pixel_avg.iloc[j] if j < len(by_pixel_avg) else None
        c.value = float(v) if pd.notnull(v) else None
        c.number_format = "0.00%"

    # Per-pixel SD row (label in column E)
    r_sd = r_sum1 + 1
    ws.cell(row=r_sd, column=1).value = "Overall SD"
    ws.cell(row=r_sd, column=1).font = bold
    c_overall_sd = ws.cell(row=r_sd, column=2)
    overall_sd = float(pd.Series(year_avg).std(ddof=0)) if len(year_avg) else None
    c_overall_sd.value = overall_sd if overall_sd is not None and pd.notnull(overall_sd) else None
    c_overall_sd.number_format = "0.00%"

    ws.cell(row=r_sd, column=col_label).value = "SD"
    ws.cell(row=r_sd, column=col_label).font = bold
    for j, pix in enumerate(pixel_order):
        c = ws.cell(row=r_sd, column=first_data_col + j)
        v = by_pixel_sd.iloc[j] if j < len(by_pixel_sd) else None
        c.value = float(v) if pd.notnull(v) else None
        c.number_format = "0.00%"

    # Min/Max/Percentiles per pixel (labels in column E)
    for offset, label, series in [
        (1, "Min", by_pixel_min),
        (2, "Max", by_pixel_max),
        (3, "90th percentile", by_pixel_p90),
        (4, "95th percentile", by_pixel_p95),
    ]:
        r = r_sd + offset
        lab = ws.cell(row=r, column=col_label)
        lab.value = label
        lab.font = bold
        for j, pix in enumerate(pixel_order):
            c = ws.cell(row=r, column=first_data_col + j)
            v = series.iloc[j] if j < len(series) else None
            c.value = float(v) if pd.notnull(v) else None
            c.number_format = "0.00%"

    # Style headers already present
    ws.cell(row=row_title, column=col_label).font = bold
 
    for rr in range(row_meta_start, row_meta_start + 7): #looping over metadata rows
        c = ws.cell(row=rr, column=col_label)
        c.font = bold
        c.alignment = left
    for cc in [2, 4, col_label]: #looping over SD, Average, Pixel ID columns
        ws.cell(row=row_header, column=cc).font = bold
        if cc == col_label:
            ws.cell(row=row_header, column=cc).alignment = left
        else:
            ws.cell(row=row_header, column=cc).alignment = center
    for j in range(len(pixel_order)):
        ws.cell(row=row_header, column=first_data_col + j).font = bold
        ws.cell(row=row_header, column=first_data_col + j).alignment = center

    last_col = max(first_data_col + len(pixel_order) - 1, col_label)
    autosize_columns(ws, 1, last_col)
    return wb

# Example usage:
# df has columns: Pixel_ID, Year, PayoutsPercent (or Percent_Payout), Attach, Detach, Area, Region, Lon, Lat
# wb = build_payouts_percent_sheet(df)
# wb.save("policy_pilot_payouts.xlsx")